import os
import csv
from openpyxl import Workbook, load_workbook

INPUT_FOLDERS = [
    "10sec,diffrenpower,10set",
    "30sec50power15set",
    "45sec diffrent powrset2",
]


def try_convert(value):
    """Preserve exact string if blank; otherwise try numeric conversion."""
    if value == "":
        return value
    # Try integer first (avoids float like 1.0 for whole numbers)
    try:
        return int(value)
    except ValueError:
        pass
    # Try float (handles scientific notation like -6.19750623094309E-10)
    try:
        return float(value)
    except ValueError:
        pass
    return value


def verify_excel_vs_csv(csv_path, excel_path):
    """
    Re-read both files and compare every cell value.
    Returns (True, "") on a perfect match, or (False, reason) on any mismatch.
    """
    # Re-read the CSV into a list-of-lists (same conversion as during write)
    csv_data = []
    with open(csv_path, "r", encoding="utf-8", errors="replace") as f:
        for row in csv.reader(f):
            csv_data.append([try_convert(v) for v in row])

    # Re-read the saved Excel
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active

    xl_row_count = ws.max_row
    csv_row_count = len(csv_data)

    if xl_row_count != csv_row_count:
        wb.close()
        return False, f"row count differs: CSV={csv_row_count}, Excel={xl_row_count}"

    for r_idx, (csv_row, xl_row) in enumerate(
            zip(csv_data, ws.iter_rows(values_only=True)), start=1):

        # Normalize Excel row: replace None with "" to match empty CSV cells
        xl_cells = [("" if v is None else v) for v in xl_row]
        csv_cells = csv_row

        # Pad the shorter side so length mismatches are caught
        max_cols = max(len(xl_cells), len(csv_cells))
        xl_cells  += [""] * (max_cols - len(xl_cells))
        csv_cells += [""] * (max_cols - len(csv_cells))

        for c_idx, (cv, xv) in enumerate(zip(csv_cells, xl_cells), start=1):
            if cv != xv:
                wb.close()
                return False, (
                    f"value mismatch at row {r_idx}, col {c_idx}: "
                    f"CSV={cv!r}  Excel={xv!r}"
                )

    wb.close()
    return True, ""


# Get root folder (same directory as this script)
folder_path = os.path.dirname(os.path.abspath(__file__))

print("Working Folder:", folder_path, "\n")
print("Input Folders:")
for folder in INPUT_FOLDERS:
    print(" ", folder)
print()

total_converted = 0
total_failed = 0
failed_files = []

# Walk only the requested input folders
for input_folder in INPUT_FOLDERS:
    start_dir = os.path.join(folder_path, input_folder)
    if not os.path.isdir(start_dir):
        failed_files.append(f"{input_folder} (folder not found)")
        total_failed += 1
        continue

    for root, dirs, files in os.walk(start_dir):
        for filename in sorted(files):
            if not filename.lower().endswith(".csv"):
                continue

            csv_path = os.path.join(root, filename)
            excel_filename = os.path.splitext(filename)[0] + ".xlsx"
            excel_path = os.path.join(root, excel_filename)
            rel_path = os.path.relpath(csv_path, folder_path)

            try:
                print(f"Converting: {rel_path}")

                wb = Workbook()
                ws = wb.active
                rows_read = 0

                # Use the csv module for correct parsing (handles quoted fields,
                # trailing commas, and embedded commas inside quoted values)
                with open(csv_path, "r", encoding="utf-8", errors="replace") as f:
                    reader = csv.reader(f)
                    for row_idx, row in enumerate(reader, start=1):
                        for col_idx, value in enumerate(row, start=1):
                            ws.cell(row=row_idx, column=col_idx, value=try_convert(value))
                        rows_read += 1

                wb.save(excel_path)

                # --- Step 2: Deep cell-by-cell verification ---
                print(f"  Verifying...", end=" ")
                match, reason = verify_excel_vs_csv(csv_path, excel_path)

                if match:
                    os.remove(csv_path)
                    print(f"MATCH  {rows_read} rows x {ws.max_column} cols  -> CSV deleted.")
                    total_converted += 1
                else:
                    print(f"MISMATCH: {reason}")
                    print(f"  CSV kept. Excel file may be corrupt — check manually.")
                    failed_files.append(f"{rel_path} ({reason})")
                    total_failed += 1

            except Exception as e:
                print(f"  ERROR: {rel_path} -> {e}")
                failed_files.append(f"{rel_path} ({e})")
                total_failed += 1

# ---- Summary ----
print(f"\n{'='*60}")
print(f"DONE: {total_converted} converted successfully, {total_failed} failed.")
if failed_files:
    print("\nFailed files:")
    for f in failed_files:
        print(f"  - {f}")
if total_converted:
    print("Matched CSV files were converted to Excel and deleted.")
else:
    print("No CSV files were converted.")
