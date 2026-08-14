import csv
import os
from dataclasses import dataclass
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook

INPUT_FOLDERS = [
    "10sec,diffrenpower,10set",
    "30sec50power15set",
    "45sec diffrent powrset2",
]


@dataclass
class Summary:
    csv_converted: int = 0
    csv_failed: int = 0
    cleaned_ok: int = 0
    cleaned_failed: int = 0
    renamed_ok: int = 0
    renamed_failed: int = 0
    deleted_ok: int = 0
    deleted_failed: int = 0


def try_convert(value: str):
    """Preserve blank strings; otherwise try int, then float, then original string."""
    if value == "":
        return value

    try:
        return int(value)
    except ValueError:
        pass

    try:
        return float(value)
    except ValueError:
        pass

    return value


def fmt(val) -> str:
    """Format numeric value: int-like values become ints, others stay as plain text."""
    try:
        f = float(val)
        return str(int(f)) if f == int(f) else str(f)
    except (TypeError, ValueError):
        return str(val)


def verify_excel_vs_csv(csv_path: str, excel_path: str) -> Tuple[bool, str]:
    """
    Re-read both files and compare every cell value.
    Returns (True, "") on a perfect match, or (False, reason) on mismatch.
    """
    csv_data = []
    with open(csv_path, "r", encoding="utf-8", errors="replace") as file_obj:
        for row in csv.reader(file_obj):
            csv_data.append([try_convert(v) for v in row])

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    worksheet = workbook.active

    xl_row_count = worksheet.max_row
    csv_row_count = len(csv_data)

    if xl_row_count != csv_row_count:
        workbook.close()
        return False, f"row count differs: CSV={csv_row_count}, Excel={xl_row_count}"

    for row_idx, (csv_row, xl_row) in enumerate(
        zip(csv_data, worksheet.iter_rows(values_only=True)), start=1
    ):
        xl_cells = [("" if v is None else v) for v in xl_row]
        csv_cells = csv_row

        max_cols = max(len(xl_cells), len(csv_cells))
        xl_cells += [""] * (max_cols - len(xl_cells))
        csv_cells += [""] * (max_cols - len(csv_cells))

        for col_idx, (csv_value, xl_value) in enumerate(zip(csv_cells, xl_cells), start=1):
            if csv_value != xl_value:
                workbook.close()
                return (
                    False,
                    f"value mismatch at row {row_idx}, col {col_idx}: "
                    f"CSV={csv_value!r} Excel={xl_value!r}",
                )

    workbook.close()
    return True, ""


def convert_csv_to_xlsx(root_folder: str, failures: List[str], summary: Summary) -> None:
    print("=" * 60)
    print("STEP 1: CSV -> XLSX conversion with verification")
    print("=" * 60)

    for input_folder in INPUT_FOLDERS:
        start_dir = os.path.join(root_folder, input_folder)
        if not os.path.isdir(start_dir):
            failures.append(f"{input_folder} (folder not found)")
            summary.csv_failed += 1
            continue

        for walk_root, _, files in os.walk(start_dir):
            for filename in sorted(files):
                if not filename.lower().endswith(".csv"):
                    continue

                csv_path = os.path.join(walk_root, filename)
                xlsx_path = os.path.join(walk_root, f"{os.path.splitext(filename)[0]}.xlsx")
                rel_path = os.path.relpath(csv_path, root_folder)

                try:
                    print(f"Converting: {rel_path}")
                    workbook = Workbook()
                    worksheet = workbook.active
                    rows_read = 0

                    with open(csv_path, "r", encoding="utf-8", errors="replace") as file_obj:
                        reader = csv.reader(file_obj)
                        for row_idx, row in enumerate(reader, start=1):
                            for col_idx, value in enumerate(row, start=1):
                                worksheet.cell(
                                    row=row_idx,
                                    column=col_idx,
                                    value=try_convert(value),
                                )
                            rows_read += 1

                    workbook.save(xlsx_path)

                    print("  Verifying...", end=" ")
                    match, reason = verify_excel_vs_csv(csv_path, xlsx_path)

                    if match:
                        os.remove(csv_path)
                        print(
                            f"MATCH {rows_read} rows x {worksheet.max_column} cols "
                            "-> CSV deleted"
                        )
                        summary.csv_converted += 1
                    else:
                        print(f"MISMATCH: {reason}")
                        failures.append(f"{rel_path} ({reason})")
                        summary.csv_failed += 1

                except Exception as ex:
                    print(f"  ERROR: {rel_path} -> {ex}")
                    failures.append(f"{rel_path} ({ex})")
                    summary.csv_failed += 1

    print(
        f"\nStep 1 result: {summary.csv_converted} converted, "
        f"{summary.csv_failed} failed.\n"
    )


def clean_xlsx_files(
    root_folder: str,
    failures: List[str],
    summary: Summary,
) -> Dict[str, str]:
    """Create cleaned workbooks from original xlsx files."""
    print("=" * 60)
    print("STEP 2: Cleaning XLSX files")
    print("=" * 60)

    cleaned_map: Dict[str, str] = {}

    for input_folder in INPUT_FOLDERS:
        start_dir = os.path.join(root_folder, input_folder)
        if not os.path.isdir(start_dir):
            failures.append(f"{input_folder} (folder not found during cleaning)")
            summary.cleaned_failed += 1
            continue

        for walk_root, _, files in os.walk(start_dir):
            for filename in sorted(files):
                if not filename.endswith(".xlsx") or filename.startswith(("cleaned_", "sec-", "~$")):
                    continue

                source_path = os.path.join(walk_root, filename)
                rel_path = os.path.relpath(source_path, root_folder)
                print(f"Cleaning: {rel_path}")

                try:
                    df = pd.read_excel(source_path, header=None)

                    # Metadata cells copied from original worksheet positions
                    a20 = df.iloc[19, 0]
                    b20 = df.iloc[19, 1]
                    a23 = df.iloc[22, 0]
                    b23 = df.iloc[22, 1]
                    a40 = df.iloc[39, 0]
                    b40 = df.iloc[39, 1]

                    # Raman data starts at row index 98 and spans through 2146 inclusive
                    raman_shift = df.iloc[98:2147, 3].reset_index(drop=True)
                    dark_sub = df.iloc[98:2147, 7].reset_index(drop=True)

                    max_len = max(len(raman_shift), 3)
                    clean_df = pd.DataFrame(index=range(max_len), columns=[0, 1, 2, 3])

                    clean_df.iloc[0, 0] = a20
                    clean_df.iloc[0, 1] = b20
                    clean_df.iloc[1, 0] = a23
                    clean_df.iloc[1, 1] = b23
                    clean_df.iloc[2, 0] = a40
                    clean_df.iloc[2, 1] = b40
                    clean_df.iloc[: len(raman_shift), 2] = raman_shift.values
                    clean_df.iloc[: len(dark_sub), 3] = dark_sub.values

                    temp_path = os.path.join(walk_root, f"cleaned_{filename}")
                    clean_df.to_excel(temp_path, index=False, header=False)
                    cleaned_map[source_path] = temp_path

                    print(f"  -> cleaned_{filename} ({len(raman_shift)} rows)")
                    summary.cleaned_ok += 1

                except Exception as ex:
                    print(f"  ERROR: {rel_path} -> {ex}")
                    failures.append(f"{rel_path} ({ex})")
                    summary.cleaned_failed += 1

    print(
        f"\nStep 2 result: {summary.cleaned_ok} cleaned, "
        f"{summary.cleaned_failed} failed.\n"
    )
    return cleaned_map


def rename_cleaned_files(
    root_folder: str,
    cleaned_map: Dict[str, str],
    failures: List[str],
    summary: Summary,
) -> Dict[str, str]:
    """Rename cleaned workbooks to sec-<b1>_power-<b3>_i-<b2>.xlsx format."""
    print("=" * 60)
    print("STEP 3: Renaming cleaned files")
    print("=" * 60)

    renamed_map: Dict[str, str] = {}
    used_names: Dict[str, Dict[str, int]] = {}

    for original_path, temp_path in cleaned_map.items():
        rel_temp = os.path.relpath(temp_path, root_folder)
        try:
            df = pd.read_excel(temp_path, header=None)

            b1 = fmt(df.iloc[0, 1])
            b2 = fmt(df.iloc[1, 1])
            b3 = fmt(df.iloc[2, 1])

            base_name = f"sec-{b1}_power-{b3}_i-{b2}"
            folder = os.path.dirname(temp_path)

            if folder not in used_names:
                used_names[folder] = {}

            if base_name not in used_names[folder]:
                used_names[folder][base_name] = 1
                new_filename = f"{base_name}.xlsx"
            else:
                used_names[folder][base_name] += 1
                suffix = used_names[folder][base_name]
                new_filename = f"{base_name}_{suffix}.xlsx"

            final_path = os.path.join(folder, new_filename)
            os.rename(temp_path, final_path)

            print(f"{rel_temp} -> {new_filename}")
            renamed_map[original_path] = final_path
            summary.renamed_ok += 1

        except Exception as ex:
            print(f"ERROR renaming {rel_temp}: {ex}")
            failures.append(f"{rel_temp} ({ex})")
            summary.renamed_failed += 1

    print(
        f"\nStep 3 result: {summary.renamed_ok} renamed, "
        f"{summary.renamed_failed} failed.\n"
    )
    return renamed_map


def verify_and_delete_originals(
    root_folder: str,
    renamed_map: Dict[str, str],
    failures: List[str],
    summary: Summary,
) -> None:
    """Delete originals only after sanity checks on renamed files pass."""
    print("=" * 60)
    print("STEP 4: Verifying renamed files and deleting originals")
    print("=" * 60)

    for original_path, final_path in renamed_map.items():
        original_rel = os.path.relpath(original_path, root_folder)
        final_rel = os.path.relpath(final_path, root_folder)

        if not os.path.exists(final_path):
            msg = f"MISSING {final_rel} - original kept: {original_rel}"
            print(msg)
            failures.append(msg)
            summary.deleted_failed += 1
            continue

        if os.path.getsize(final_path) == 0:
            msg = f"EMPTY {final_rel} - original kept: {original_rel}"
            print(msg)
            failures.append(msg)
            summary.deleted_failed += 1
            continue

        try:
            df_check = pd.read_excel(final_path, header=None)
            if len(df_check) < 3:
                raise ValueError(f"only {len(df_check)} rows found")
        except Exception as ex:
            msg = f"CORRUPT {final_rel} ({ex}) - original kept: {original_rel}"
            print(msg)
            failures.append(msg)
            summary.deleted_failed += 1
            continue

        os.remove(original_path)
        print(f"OK {final_rel} -> deleted {os.path.basename(original_path)}")
        summary.deleted_ok += 1

    print(
        f"\nStep 4 result: {summary.deleted_ok} originals removed, "
        f"{summary.deleted_failed} kept.\n"
    )


def main() -> None:
    root_folder = os.path.dirname(os.path.abspath(__file__))
    failures: List[str] = []
    summary = Summary()

    print("Working folder:", root_folder)
    print("Input folders:")
    for input_folder in INPUT_FOLDERS:
        print(" ", input_folder)
    print()

    convert_csv_to_xlsx(root_folder, failures, summary)
    cleaned_map = clean_xlsx_files(root_folder, failures, summary)
    renamed_map = rename_cleaned_files(root_folder, cleaned_map, failures, summary)
    verify_and_delete_originals(root_folder, renamed_map, failures, summary)

    print("=" * 60)
    print("FINAL SUMMARY")
    print("=" * 60)
    print(f"STEP 1 CSV->XLSX: {summary.csv_converted} ok, {summary.csv_failed} failed")
    print(f"STEP 2 Cleaned  : {summary.cleaned_ok} ok, {summary.cleaned_failed} failed")
    print(f"STEP 3 Renamed  : {summary.renamed_ok} ok, {summary.renamed_failed} failed")
    print(f"STEP 4 Deleted  : {summary.deleted_ok} ok, {summary.deleted_failed} failed")

    if failures:
        print("\nIssues to review:")
        for item in failures:
            print(f"- {item}")


if __name__ == "__main__":
    main()
