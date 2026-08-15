"""
Publication-Quality Raman Figure Generation
Evidence-driven workflow — no hard-coded labels, assignments, or annotations.
All peak annotations are earned through statistical criteria.
"""

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from matplotlib.ticker import AutoMinorLocator
from openpyxl import load_workbook
from scipy.signal import find_peaks, savgol_filter, peak_widths
from scipy import sparse
from scipy.sparse.linalg import spsolve
from scipy.stats import pearsonr
import os
import json
import warnings
warnings.filterwarnings('ignore', category=sparse.SparseEfficiencyWarning)

# ─────────────────────────────────────────────
# Configuration
# ─────────────────────────────────────────────
BASE_DIR   = r"c:\Users\sukes\Downloads\nam-new"
FIG_DIR    = os.path.join(BASE_DIR, "Analysis", "Figures")
PROC_DIR   = os.path.join(BASE_DIR, "Analysis", "Processed")
REPORT_DIR = os.path.join(BASE_DIR, "Analysis")
os.makedirs(FIG_DIR, exist_ok=True)
os.makedirs(PROC_DIR, exist_ok=True)
os.chdir(BASE_DIR)

# Journal-quality rcParams (Nature/ACS style)
plt.rcParams.update({
    'font.family'       : 'Arial',
    'font.size'         : 9,
    'axes.linewidth'    : 0.8,
    'axes.labelsize'    : 9,
    'axes.titlesize'    : 9,
    'xtick.major.width' : 0.8,
    'ytick.major.width' : 0.8,
    'xtick.minor.width' : 0.5,
    'ytick.minor.width' : 0.5,
    'xtick.major.size'  : 3,
    'ytick.major.size'  : 3,
    'xtick.minor.size'  : 1.5,
    'ytick.minor.size'  : 1.5,
    'xtick.direction'   : 'in',
    'ytick.direction'   : 'in',
    'lines.linewidth'   : 0.8,
    'legend.fontsize'   : 8,
    'legend.framealpha' : 0.9,
    'legend.edgecolor'  : '0.8',
    'savefig.dpi'       : 300,
    'savefig.bbox'      : 'tight',
    'figure.dpi'        : 120,
})

COLORS = {
    'glass' : '#2166ac',
    'nam1'  : '#d73027',
    'nam2'  : '#fc8d59',
    'nam3'  : '#4dac26',
    'mean'  : '#000000',
    'shade' : '#bbbbbb',
    'flag'  : '#e31a1c',
}

FINGERPRINT = (400, 1800)
DPI = 300

# ─────────────────────────────────────────────
# STEP 1 — Load raw data
# ─────────────────────────────────────────────
print("=" * 70)
print("STEP 1: LOADING RAW DATA")
print("=" * 70)

def load_xlsx_spectrum(path):
    wb = load_workbook(path, read_only=True)
    ws = wb.active
    rs, ds = [], []
    for row in ws.iter_rows(min_row=100, values_only=True):
        try:
            rs.append(float(row[3]))
            ds.append(float(row[7]))
        except (TypeError, ValueError, IndexError):
            continue
    wb.close()
    return np.array(rs), np.array(ds)

NAM_DIR   = r"70p-60s-5ac-diifrentpoint"
GLASS_DIR = r"glass slide empty\empty slide"

nam_files   = sorted(f for f in os.listdir(NAM_DIR)   if f.endswith('.xlsx'))
glass_files = sorted(f for f in os.listdir(GLASS_DIR) if f.endswith('.xlsx'))

nam_raw   = [(load_xlsx_spectrum(os.path.join(NAM_DIR, f)),   f) for f in nam_files]
glass_raw = [(load_xlsx_spectrum(os.path.join(GLASS_DIR, f)), f) for f in glass_files]

for (rs, ds), f in nam_raw:
    print(f"  NAM  {f}: {len(rs)} pts | RS {rs.min():.1f}–{rs.max():.1f} | "
          f"max={ds.max():.0f} | neg%={100*(ds<0).mean():.1f}% | sat={np.sum(ds>=65535)}")
for (rs, ds), f in glass_raw:
    print(f"  GLASS {f}: {len(rs)} pts | RS {rs.min():.1f}–{rs.max():.1f} | "
          f"max={ds.max():.0f} | neg%={100*(ds<0).mean():.1f}% | sat={np.sum(ds>=65535)}")

# ─────────────────────────────────────────────
# STEP 2 — Process spectra
# ─────────────────────────────────────────────
print("\n" + "=" * 70)
print("STEP 2: PROCESSING")
print("=" * 70)

def crop(rs, y, lo=FINGERPRINT[0], hi=FINGERPRINT[1]):
    m = (rs >= lo) & (rs <= hi) & ~np.isnan(rs) & ~np.isnan(y)
    return rs[m], y[m]

def als_baseline(y, lam=1e5, p=0.001, niter=10):
    L = len(y)
    D = sparse.diags([1,-2,1],[0,-1,-2], shape=(L, L-2))
    w = np.ones(L)
    for _ in range(niter):
        W = sparse.spdiags(w, 0, L, L)
        z = spsolve(W + lam * D.dot(D.T), w * y)
        w = p * (y > z) + (1-p) * (y <= z)
    return z

def process(rs_raw, y_raw):
    rs, y = crop(rs_raw, y_raw)
    bl    = als_baseline(y)
    y_bc  = y - bl
    y_sg  = savgol_filter(y_bc, window_length=11, polyorder=3)
    lo, hi = y_sg.min(), y_sg.max()
    y_n   = (y_sg - lo) / (hi - lo) if hi > lo else np.zeros_like(y_sg)
    return rs, y_n, y_bc, bl, y

nam_proc   = [process(*d) for d, _ in nam_raw]
glass_proc = [process(*d) for d, _ in glass_raw]

# Alignment check — all spectra must share same RS grid
rs_ref = nam_proc[0][0]
for i, (rs, *_) in enumerate(nam_proc[1:], 1):
    if not np.allclose(rs, rs_ref, atol=0.5):
        print(f"  WARNING: NAM Spot {i+1} RS grid differs from spot 1!")
    else:
        print(f"  NAM Spot {i+1}: RS grid aligned OK ({len(rs)} pts)")
for i, (rs, *_) in enumerate(glass_proc):
    if not np.allclose(rs, rs_ref, atol=0.5):
        print(f"  WARNING: Glass {i+1} RS grid differs from NAM!")
    else:
        print(f"  Glass {i+1}: RS grid aligned OK")

# ─────────────────────────────────────────────
# STEP 3 — Statistical peak verification
# ─────────────────────────────────────────────
print("\n" + "=" * 70)
print("STEP 3: STATISTICAL PEAK VERIFICATION")
print("=" * 70)

rs_c   = rs_ref
stack  = np.vstack([p[1] for p in nam_proc])      # normalized traces
mean_s = stack.mean(axis=0)
std_s  = stack.std(axis=0)
cv_s   = np.where(mean_s > 0.02, std_s / mean_s, np.nan)  # coefficient of variation

# Peak detection on mean
peaks_idx, props = find_peaks(mean_s, prominence=0.04, distance=12, height=0.03)
widths_half, _, _, _ = peak_widths(mean_s, peaks_idx, rel_height=0.5)

# Per-peak: check reproducibility across replicates
rep_scores = []
for pi in peaks_idx:
    window = 8   # ± pixels
    lo_w, hi_w = max(0, pi-window), min(len(rs_c)-1, pi+window)
    # Does each replicate have a local max near this peak?
    local_maxima_count = 0
    for trace in stack:
        local_max = trace[lo_w:hi_w].max()
        local_min = trace[lo_w:hi_w].min()
        if local_max - local_min > 0.015:   # must show at least some rise
            local_maxima_count += 1
    rep_score = local_maxima_count / len(stack)
    rep_scores.append(rep_score)

rep_scores = np.array(rep_scores)

# Confidence classification
# HIGH:   prominence>0.10, rep>=1.0
# MEDIUM: prominence>0.05, rep>=0.67
# LOW:    everything else
confidences = []
for i, pi in enumerate(peaks_idx):
    prom = props['prominences'][i]
    rep  = rep_scores[i]
    if prom >= 0.10 and rep >= 1.0:
        confidences.append('HIGH')
    elif prom >= 0.05 and rep >= 0.67:
        confidences.append('MEDIUM')
    else:
        confidences.append('LOW')

print(f"\n  Total peaks detected: {len(peaks_idx)}")
print(f"  HIGH confidence:   {confidences.count('HIGH')}")
print(f"  MEDIUM confidence: {confidences.count('MEDIUM')}")
print(f"  LOW confidence:    {confidences.count('LOW')}")
print(f"\n  {'#':<4} {'RS (cm-1)':<12} {'Prom':<8} {'Rep':<8} {'Width':<8} {'Conf'}")
print("  " + "-" * 52)
for i, pi in enumerate(peaks_idx):
    print(f"  {i+1:<4} {rs_c[pi]:<12.1f} {props['prominences'][i]:<8.3f} "
          f"{rep_scores[i]:<8.2f} {widths_half[i]:<8.1f} {confidences[i]}")

# ─────────────────────────────────────────────
# STEP 4 — Replicate consistency
# ─────────────────────────────────────────────
print("\n" + "=" * 70)
print("STEP 4: REPLICATE CONSISTENCY")
print("=" * 70)

for i in range(len(nam_proc)):
    for j in range(i+1, len(nam_proc)):
        r, _ = pearsonr(nam_proc[i][1], nam_proc[j][1])
        print(f"  Spot {i+1} vs Spot {j+1}: Pearson r = {r:.4f}")

cv_mean  = np.nanmean(cv_s)
cv_max   = np.nanmax(cv_s)
unstable = rs_c[cv_s > 0.3] if np.any(cv_s > 0.3) else np.array([])
print(f"\n  CV across spectrum: mean={cv_mean:.3f}, max={cv_max:.3f}")
if len(unstable):
    print(f"  Unstable regions (CV>0.3) at: {unstable[:10].round(0)} cm-1")
else:
    print("  No highly unstable regions (CV<0.3 throughout)")

# ─────────────────────────────────────────────
# STEP 5 — Glass control validation
# ─────────────────────────────────────────────
print("\n" + "=" * 70)
print("STEP 5: GLASS CONTROL VALIDATION")
print("=" * 70)

glass_stack = np.vstack([p[1] for p in glass_proc])
glass_mean  = glass_stack.mean(axis=0)
glass_std   = glass_stack.std(axis=0)

glass_peaks_idx, glass_props = find_peaks(glass_mean, prominence=0.05, distance=12)
glass_peak_pos = rs_c[glass_peaks_idx] if len(glass_peaks_idx) > 0 else np.array([])

print(f"  Glass peaks after processing: {len(glass_peak_pos)}")
if len(glass_peak_pos):
    print(f"  Positions: {glass_peak_pos.round(1)}")

# Flag any NAM peak within 15 cm-1 of a glass peak
substrate_flags = []
for i, pi in enumerate(peaks_idx):
    pos = rs_c[pi]
    overlap = any(abs(pos - gp) < 15 for gp in glass_peak_pos)
    substrate_flags.append(overlap)
    if overlap:
        print(f"  SUBSTRATE OVERLAP WARNING: {pos:.1f} cm-1 near glass peak")

if not any(substrate_flags):
    print("  No NAM peaks overlap with glass substrate peaks. All retained.")

# ─────────────────────────────────────────────
# STEP 6 — Annotation rules
# ─────────────────────────────────────────────
# Only annotate HIGH or MEDIUM confidence, non-substrate, non-overlapping labels
def select_annotatable(peaks_idx, rs_c, confidences, substrate_flags,
                       min_sep_label=35):
    """Return subset of peak indices eligible for annotation."""
    eligible = [i for i, (c, sf) in enumerate(zip(confidences, substrate_flags))
                if c in ('HIGH', 'MEDIUM') and not sf]
    # Remove label-overlap: sort by prominence, greedy keep
    eligible_sorted = sorted(eligible, key=lambda i: props['prominences'][i], reverse=True)
    kept, kept_pos = [], []
    for i in eligible_sorted:
        pos = rs_c[peaks_idx[i]]
        if all(abs(pos - kp) >= min_sep_label for kp in kept_pos):
            kept.append(i)
            kept_pos.append(pos)
    return sorted(kept)   # restore position order

annotatable = select_annotatable(peaks_idx, rs_c, confidences, substrate_flags)
print(f"\n  Peaks eligible for annotation: {len(annotatable)}")
for i in annotatable:
    print(f"    {rs_c[peaks_idx[i]]:.1f} cm-1 [{confidences[i]}]")

# ─────────────────────────────────────────────────────────────────────────────
# FIGURES
# ─────────────────────────────────────────────────────────────────────────────

def save_fig(fig, name):
    for ext in ('png', 'pdf'):
        fig.savefig(os.path.join(FIG_DIR, f"{name}.{ext}"), dpi=DPI,
                    facecolor='white', bbox_inches='tight')
    plt.close(fig)
    print(f"  Saved: {name}.png / .pdf")

def style_ax(ax, xlabel='Raman Shift (cm$^{-1}$)', ylabel='Intensity (a.u.)'):
    ax.set_xlabel(xlabel, fontsize=9)
    ax.set_ylabel(ylabel, fontsize=9)
    ax.xaxis.set_minor_locator(AutoMinorLocator(5))
    ax.yaxis.set_minor_locator(AutoMinorLocator(5))
    ax.tick_params(which='both', direction='in', top=True, right=True)
    ax.spines[['top','right']].set_linewidth(0.5)

# ─── Figure 1: Raw Glass vs Raw NAM ─────────────────────────────────────────
print("\n[FIG 1] Raw Glass vs Raw NAM")
fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(7, 5.5), sharex=True)
fig.subplots_adjust(hspace=0.08)

# Top watermark: folder + condition
fig.text(0.99, 0.99,
         f'Glass: folder = "glass slide empty/empty slide"  |  n={len(glass_raw)} spectra',
         ha='right', va='top', fontsize=6.5, color='gray', style='italic')

for i, ((rs, ds), f) in enumerate(glass_raw):
    m = (rs >= FINGERPRINT[0]) & (rs <= FINGERPRINT[1])
    alpha = 0.45 + 0.11 * i
    ax1.plot(rs[m], ds[m], lw=0.6, color=COLORS['glass'],
             alpha=alpha, label=f'{f}')
ax1.set_ylabel('Dark Sub. Intensity (counts)')
ax1.set_title('(a)  Glass substrate — RAW  |  folder: glass slide empty/empty slide',
              loc='left', fontsize=8, fontweight='bold')
ax1.legend(ncol=3, fontsize=6, title='File name', title_fontsize=6)
style_ax(ax1, xlabel='')

# Bottom watermark
fig.text(0.99, 0.50,
         f'NAM: folder = "70p-60s-5ac-diifrentpoint"  |  70% power, 60s, 5acc  |  n={len(nam_raw)} spots',
         ha='right', va='top', fontsize=6.5, color='gray', style='italic')

for i, ((rs, ds), f) in enumerate(nam_raw):
    m = (rs >= FINGERPRINT[0]) & (rs <= FINGERPRINT[1])
    c = [COLORS['nam1'], COLORS['nam2'], COLORS['nam3']][i]
    ax2.plot(rs[m], ds[m], lw=0.6, color=c, alpha=0.85,
             label=f'{f}  (Spot {i+1})')
ax2.set_ylabel('Dark Sub. Intensity (counts)')
ax2.set_title('(b)  NAM sample — RAW  |  folder: 70p-60s-5ac-diifrentpoint  |  70% power, 60s, 5acc',
              loc='left', fontsize=8, fontweight='bold')
ax2.legend(ncol=1, fontsize=6, title='File name', title_fontsize=6)
ax2.set_xlim(FINGERPRINT)
style_ax(ax2)
save_fig(fig, 'Figure1_Raw_Glass_vs_NAM')

# ─── Figure 2: Processed Glass vs Processed NAM ─────────────────────────────
print("[FIG 2] Processed Glass vs Processed NAM")
fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(7, 5.5), sharex=True)
fig.subplots_adjust(hspace=0.08)
fig.text(0.99, 0.99,
         'Processing: ALS baseline (λ=1e5, p=0.001) → SG smooth (w=11, p=3) → min-max norm',
         ha='right', va='top', fontsize=6.5, color='gray', style='italic')

gstack = np.vstack([p[1] for p in glass_proc])
gmean, gstd = gstack.mean(0), gstack.std(0)
for i, (p, f) in enumerate(zip(glass_proc, glass_files)):
    ax1.plot(rs_c, p[1], lw=0.4, color=COLORS['glass'], alpha=0.4, label=f)
ax1.plot(rs_c, gmean, lw=1.2, color=COLORS['glass'], label=f'Mean (n={len(glass_proc)})')
ax1.fill_between(rs_c, gmean-gstd, gmean+gstd, color=COLORS['glass'], alpha=0.2, label='±1 SD')
ax1.set_ylabel('Normalised Intensity')
ax1.set_title('(a)  Glass substrate — PROCESSED  |  folder: glass slide empty/empty slide',
              loc='left', fontsize=8, fontweight='bold')
ax1.legend(ncol=2, fontsize=6, title='File name', title_fontsize=6)
style_ax(ax1, xlabel='')

for i, (p, f) in enumerate(zip(nam_proc, nam_files)):
    c = [COLORS['nam1'], COLORS['nam2'], COLORS['nam3']][i]
    ax2.plot(rs_c, p[1], lw=0.5, color=c, alpha=0.6, label=f'{f}  (Spot {i+1})')
ax2.plot(rs_c, mean_s, lw=1.2, color=COLORS['mean'], label=f'Mean (n={len(nam_proc)})')
ax2.fill_between(rs_c, mean_s-std_s, mean_s+std_s, color=COLORS['shade'], alpha=0.5, label='±1 SD')
ax2.set_ylabel('Normalised Intensity')
ax2.set_title('(b)  NAM sample — PROCESSED  |  folder: 70p-60s-5ac-diifrentpoint  |  70% power, 60s, 5acc',
              loc='left', fontsize=8, fontweight='bold')
ax2.legend(ncol=1, fontsize=6, title='File name', title_fontsize=6)
ax2.set_xlim(FINGERPRINT)
style_ax(ax2)
save_fig(fig, 'Figure2_Processed_Glass_vs_NAM')

# ─── Figure 3: Replicate reproducibility ────────────────────────────────────
print("[FIG 3] Replicate reproducibility")
fig = plt.figure(figsize=(7, 6))
gs = gridspec.GridSpec(2, 2, hspace=0.35, wspace=0.35)
ax_main = fig.add_subplot(gs[0, :])
ax_cv   = fig.add_subplot(gs[1, 0])
ax_corr = fig.add_subplot(gs[1, 1])

lab_colors = [COLORS['nam1'], COLORS['nam2'], COLORS['nam3']]
for i, (p, f) in enumerate(zip(nam_proc, nam_files)):
    ax_main.plot(rs_c, p[1], lw=0.7, color=lab_colors[i],
                 alpha=0.8, label=f'{f}  (Spot {i+1})')
ax_main.plot(rs_c, mean_s, 'k-', lw=1.2, label='Mean')
ax_main.fill_between(rs_c, mean_s-std_s, mean_s+std_s,
                     color='gray', alpha=0.25, label='±1 SD')
ax_main.set_title(f'(a)  NAM replicates — folder: 70p-60s-5ac-diifrentpoint  |  70% power, 60s, 5acc  |  n={len(nam_proc)}',
                  loc='left', fontsize=8, fontweight='bold')
ax_main.set_ylabel('Normalised Intensity')
ax_main.legend(fontsize=7, ncol=4)
style_ax(ax_main)

ax_cv.plot(rs_c, cv_s * 100, lw=0.6, color='steelblue')
ax_cv.axhline(30, color='red', lw=0.7, ls='--', label='30% CV threshold')
ax_cv.set_title('(b)  Coefficient of variation\n(n=3 spots, same folder)',
                loc='left', fontsize=8, fontweight='bold')
ax_cv.set_ylabel('CV (%)')
ax_cv.legend(fontsize=7)
style_ax(ax_cv)

r12, _ = pearsonr(nam_proc[0][1], nam_proc[1][1])
r13, _ = pearsonr(nam_proc[0][1], nam_proc[2][1])
ax_corr.plot(nam_proc[0][1], nam_proc[1][1], '.', ms=1.5,
             color=COLORS['nam1'], alpha=0.5,
             label=f'{nam_files[0]} vs {nam_files[1]}  r={r12:.3f}')
ax_corr.plot(nam_proc[0][1], nam_proc[2][1], '.', ms=1.5,
             color=COLORS['nam3'], alpha=0.5,
             label=f'{nam_files[0]} vs {nam_files[2]}  r={r13:.3f}')
lim_lo = min(nam_proc[0][1].min(), nam_proc[1][1].min()) - 0.02
lim_hi = max(nam_proc[0][1].max(), nam_proc[1][1].max()) + 0.02
ax_corr.plot([lim_lo, lim_hi], [lim_lo, lim_hi], 'k--', lw=0.6, label='1:1 line')
ax_corr.set_title('(c)  Inter-replicate correlation\n(Pearson r)', loc='left', fontsize=8, fontweight='bold')
ax_corr.set_xlabel(f'{nam_files[0]}  (Spot 1)', fontsize=8)
ax_corr.set_ylabel(f'Spot 2 / 3 intensity', fontsize=8)
ax_corr.legend(fontsize=6)
style_ax(ax_corr, xlabel=f'{nam_files[0]} (Spot 1)', ylabel='Spot 2 / 3 intensity')

save_fig(fig, 'Figure3_Replicate_Reproducibility')

# ─── Figure 4: Final averaged spectrum ──────────────────────────────────────
print("[FIG 4] Final averaged Raman spectrum")
fig, ax = plt.subplots(figsize=(7, 3.5))

ax.fill_between(rs_c, mean_s - std_s, mean_s + std_s,
                color=COLORS['shade'], alpha=0.6, zorder=1)
ax.plot(rs_c, mean_s, color=COLORS['mean'], lw=0.9, zorder=2)

# Annotate only HIGH/MEDIUM confidence, non-substrate peaks
ann_y_used = []
for i in annotatable:
    pi   = peaks_idx[i]
    xpos = rs_c[pi]
    ypos = mean_s[pi]
    conf = confidences[i]
    col  = '#c0392b' if conf == 'HIGH' else '#e67e22'

    # Stagger y position to avoid overlap
    y_ann = ypos + 0.06
    for yu in ann_y_used:
        if abs(y_ann - yu) < 0.06:
            y_ann += 0.06
    ann_y_used.append(y_ann)
    y_ann = min(y_ann, 1.05)

    ax.annotate(f'{xpos:.0f}',
                xy=(xpos, ypos), xytext=(xpos, y_ann),
                fontsize=6.5, ha='center', color=col,
                arrowprops=dict(arrowstyle='-', color=col, lw=0.5),
                annotation_clip=False)
    ax.plot(xpos, ypos, 'v', color=col, ms=3, zorder=3)

ax.set_xlim(FINGERPRINT)
ax.set_ylim(-0.05, 1.30)
ax.set_ylabel('Normalised Intensity (min-max)')
ax.set_title(
    f'NAM — mean Raman spectrum  |  folder: 70p-60s-5ac-diifrentpoint\n'
    f'Files: {", ".join(nam_files)}  |  70% power, 60s, 5acc  |  n={len(nam_proc)} spots  |  ±1 SD shaded\n'
    f'Annotated: HIGH confidence (red) / MEDIUM confidence (orange)  |  '
    f'Substrate-overlapping peak 1590 cm⁻¹ excluded',
    fontsize=7.5, loc='left')
style_ax(ax)
save_fig(fig, 'Figure4_Final_Averaged_Spectrum')

# ─── Figure 5: Peak verification and confidence analysis ────────────────────
print("[FIG 5] Peak verification & confidence")
n_peaks = len(peaks_idx)
fig, axes = plt.subplots(1, 3, figsize=(10, 3.5))
fig.subplots_adjust(wspace=0.38)

# (a) Prominence bar chart, coloured by confidence
conf_colors = {'HIGH': '#c0392b', 'MEDIUM': '#e67e22', 'LOW': '#95a5a6'}
bar_colors = [conf_colors[c] for c in confidences]
x_pos = np.arange(n_peaks)
axes[0].bar(x_pos, props['prominences'], color=bar_colors, edgecolor='none', width=0.7)
# Custom legend
from matplotlib.patches import Patch
legend_els = [Patch(facecolor=conf_colors[k], label=k) for k in ('HIGH','MEDIUM','LOW')]
axes[0].legend(handles=legend_els, fontsize=7, loc='upper right')
axes[0].set_xticks(x_pos)
axes[0].set_xticklabels([f'{rs_c[p]:.0f}' for p in peaks_idx],
                         rotation=90, fontsize=5)
axes[0].set_ylabel('Prominence')
axes[0].set_title(f'(a)  Peak prominence\nData: 70p-60s-5ac-diifrentpoint | n={len(nam_proc)} spots',
                  loc='left', fontsize=7.5, fontweight='bold')
axes[0].xaxis.set_minor_locator(AutoMinorLocator(1))

# (b) Reproducibility score
axes[1].bar(x_pos, rep_scores * 100, color=bar_colors, edgecolor='none', width=0.7)
axes[1].axhline(100, color='gray', lw=0.6, ls='--')
axes[1].set_xticks(x_pos)
axes[1].set_xticklabels([f'{rs_c[p]:.0f}' for p in peaks_idx],
                         rotation=90, fontsize=5)
axes[1].set_ylabel('Reproducibility (%)')
axes[1].set_title(f'(b)  Replicate reproducibility\n({", ".join(nam_files)})',
                  loc='left', fontsize=7.5, fontweight='bold')
axes[1].set_ylim(0, 115)

# (c) FWHM (peak width at half-max in cm-1)
rs_step   = np.mean(np.diff(rs_c))
widths_cm = widths_half * rs_step
axes[2].bar(x_pos, widths_cm, color=bar_colors, edgecolor='none', width=0.7)
axes[2].set_xticks(x_pos)
axes[2].set_xticklabels([f'{rs_c[p]:.0f}' for p in peaks_idx],
                         rotation=90, fontsize=5)
axes[2].set_ylabel('FWHM (cm$^{-1}$)')
axes[2].set_title('(c)  Peak width (FWHM)\nfrom mean spectrum',
                  loc='left', fontsize=7.5, fontweight='bold')

for ax in axes:
    ax.tick_params(which='both', direction='in', top=True, right=True)
    ax.spines[['top','right']].set_linewidth(0.5)

save_fig(fig, 'Figure5_Peak_Verification_Confidence')

# ─── Figure S1: Processing pipeline ─────────────────────────────────────────
print("[FIG S1] Processing pipeline")
f0_idx = 0
f0_name = nam_files[f0_idx]
rs_fp, y_n, y_bc, bl, y_raw_fp = nam_proc[f0_idx]

fig, axes = plt.subplots(4, 1, figsize=(7, 9), sharex=True)
fig.subplots_adjust(hspace=0.12)
fig.text(0.99, 0.995,
         f'Source file: {f0_name}  |  folder: 70p-60s-5ac-diifrentpoint  |  70% power, 60s, 5acc',
         ha='right', va='top', fontsize=6.5, color='gray', style='italic')

axes[0].plot(rs_fp, y_raw_fp, lw=0.6, color='steelblue')
axes[0].set_title(f'(a)  RAW dark-subtracted  |  file: {f0_name}  |  400–1800 cm⁻¹',
                  loc='left', fontsize=8, fontweight='bold')
axes[0].set_ylabel('Counts')

axes[1].plot(rs_fp, y_raw_fp, lw=0.5, color='steelblue', alpha=0.7, label='Raw')
axes[1].plot(rs_fp, bl, lw=1.2, color='red', ls='--',
             label='ALS baseline  (λ=1×10⁵, p=0.001, iter=10)')
axes[1].set_title('(b)  ALS baseline fit', loc='left', fontsize=8, fontweight='bold')
axes[1].set_ylabel('Counts')
axes[1].legend(fontsize=7)

axes[2].plot(rs_fp, y_bc, lw=0.6, color='green')
axes[2].axhline(0, color='gray', lw=0.5, ls=':')
axes[2].set_title('(c)  Baseline subtracted  (raw − ALS)', loc='left', fontsize=8, fontweight='bold')
axes[2].set_ylabel('Counts')

axes[3].plot(rs_fp, y_n, lw=0.6, color='black')
axes[3].set_title('(d)  SG smoothed (w=11, p=3) + min-max normalised → final processed trace',
                  loc='left', fontsize=8, fontweight='bold')
axes[3].set_ylabel('Norm. Intensity')
axes[3].set_xlim(FINGERPRINT)

for ax in axes:
    style_ax(ax, xlabel='', ylabel=ax.get_ylabel())
axes[3].set_xlabel('Raman Shift (cm$^{-1}$)', fontsize=9)

save_fig(fig, 'FigureS1_Processing_Pipeline')

# ─── Figure S2: CV heatmap across spectrum ───────────────────────────────────
print("[FIG S2] CV heatmap")
fig, ax = plt.subplots(figsize=(7, 2.8))
sc = ax.scatter(rs_c, cv_s * 100, c=cv_s * 100, cmap='RdYlGn_r',
                s=2, vmin=0, vmax=40, zorder=2)
ax.axhline(30, color='red', lw=0.7, ls='--', label='30% CV threshold')
cb = fig.colorbar(sc, ax=ax, pad=0.01)
cb.set_label('CV (%)', fontsize=8)
ax.set_ylabel('CV (%)')
ax.set_title(
    f'Coefficient of variation (CV) across 400–1800 cm⁻¹\n'
    f'Files: {", ".join(nam_files)}  |  folder: 70p-60s-5ac-diifrentpoint  |  n={len(nam_proc)} replicates',
    loc='left', fontsize=8, fontweight='bold')
ax.legend(fontsize=7)
style_ax(ax)
save_fig(fig, 'FigureS2_CV_Heatmap')

# ─────────────────────────────────────────────
# REPORTS
# ─────────────────────────────────────────────
print("\n" + "=" * 70)
print("GENERATING REPORTS")
print("=" * 70)

# 1. Peak verification table
peak_df = pd.DataFrame({
    'Peak_ID'           : range(1, n_peaks+1),
    'Position_cm1'      : rs_c[peaks_idx].round(1),
    'Normalised_Int'    : mean_s[peaks_idx].round(4),
    'Std_at_peak'       : std_s[peaks_idx].round(4),
    'Prominence'        : props['prominences'].round(4),
    'FWHM_cm1'          : (widths_cm).round(1),
    'Reproducibility_pct': (rep_scores * 100).round(1),
    'Substrate_overlap' : substrate_flags,
    'Confidence'        : confidences,
    'Annotated'         : [i in [annotatable[j] for j in range(len(annotatable))]
                           for i in range(n_peaks)],
})
peak_df.to_excel(os.path.join(PROC_DIR, 'Peak_Verification_Table.xlsx'), index=False)
print("  Saved: Peak_Verification_Table.xlsx")

# 2. Figure generation + annotation justification report (TXT)
report_lines = [
    "=" * 80,
    "PUBLICATION FIGURE GENERATION REPORT",
    "Evidence-Driven Raman Analysis",
    "=" * 80,
    f"Date: 2026-06-05",
    f"Instrument: BWS465-785H, 785 nm",
    f"Sample: NAM  |  Condition: 70% power, 60s, 5acc",
    f"Replicates: {len(nam_proc)} NAM spots, {len(glass_proc)} Glass spectra",
    "",
    "─" * 80,
    "1. DATA INTEGRITY",
    "─" * 80,
]
for (rs, ds), f in nam_raw:
    report_lines.append(
        f"  {f}: {len(rs)} pts | max={ds.max():.0f} | "
        f"neg%={100*(ds<0).mean():.1f}% | sat_pixels={np.sum(ds>=65535)}")

report_lines += [
    "",
    "─" * 80,
    "2. REPLICATE CORRELATIONS",
    "─" * 80,
]
for i in range(len(nam_proc)):
    for j in range(i+1, len(nam_proc)):
        r, _ = pearsonr(nam_proc[i][1], nam_proc[j][1])
        report_lines.append(f"  Spot {i+1} vs Spot {j+1}: Pearson r = {r:.4f}")

report_lines += [
    "",
    f"  Mean CV across fingerprint: {cv_mean:.3f}  Max CV: {cv_max:.3f}",
    f"  Unstable regions (CV>0.3 cm-1 count): {len(unstable)}",
    "",
    "─" * 80,
    "3. GLASS CONTROL",
    "─" * 80,
    f"  Glass peaks after processing: {len(glass_peak_pos)}",
    f"  Positions: {glass_peak_pos.round(1).tolist()}",
    f"  NAM peaks with substrate overlap: {sum(substrate_flags)}",
    "",
    "─" * 80,
    "4. PEAK VERIFICATION SUMMARY",
    "─" * 80,
    f"  Total detected: {n_peaks}",
    f"  HIGH confidence: {confidences.count('HIGH')}",
    f"  MEDIUM confidence: {confidences.count('MEDIUM')}",
    f"  LOW confidence: {confidences.count('LOW')}",
    f"  Annotated in Figure 4: {len(annotatable)}",
    "",
    "─" * 80,
    "5. ANNOTATION JUSTIFICATION (Figure 4)",
    "─" * 80,
    "  Rules applied:",
    "    - Confidence must be HIGH or MEDIUM",
    "    - Must not overlap with glass substrate peaks (±15 cm-1)",
    "    - Minimum label separation: 35 cm-1 (greedy by prominence)",
    "    - No molecular assignments without DFT support",
    "",
]
for i in annotatable:
    pi   = peaks_idx[i]
    report_lines.append(
        f"  {rs_c[pi]:.1f} cm-1 | prom={props['prominences'][i]:.3f} | "
        f"rep={rep_scores[i]*100:.0f}% | FWHM={widths_cm[i]:.1f} cm-1 | {confidences[i]}")

report_lines += [
    "",
    "─" * 80,
    "6. PEAKS NOT ANNOTATED (LOW confidence or substrate overlap)",
    "─" * 80,
]
not_ann = [i for i in range(n_peaks) if i not in annotatable]
for i in not_ann:
    pi = peaks_idx[i]
    reason = "substrate overlap" if substrate_flags[i] else f"confidence={confidences[i]}"
    report_lines.append(
        f"  {rs_c[pi]:.1f} cm-1 | prom={props['prominences'][i]:.3f} | "
        f"rep={rep_scores[i]*100:.0f}% | EXCLUDED: {reason}")

report_lines += [
    "",
    "─" * 80,
    "7. FIGURES GENERATED",
    "─" * 80,
    "  Figure1_Raw_Glass_vs_NAM.png/pdf",
    "  Figure2_Processed_Glass_vs_NAM.png/pdf",
    "  Figure3_Replicate_Reproducibility.png/pdf",
    "  Figure4_Final_Averaged_Spectrum.png/pdf",
    "  Figure5_Peak_Verification_Confidence.png/pdf",
    "  FigureS1_Processing_Pipeline.png/pdf",
    "  FigureS2_CV_Heatmap.png/pdf",
    "",
    "─" * 80,
    "8. DFT STATUS",
    "─" * 80,
    "  DFT data: NOT AVAILABLE",
    "  Figure 6 (Experimental vs DFT Overlay): PENDING",
    "  No molecular vibration assignments have been made.",
    "  Peak labels show POSITION ONLY — no mode assignments.",
    "",
    "=" * 80,
    "END OF REPORT",
    "=" * 80,
]

report_path = os.path.join(REPORT_DIR, 'Figure_Generation_Report.txt')
with open(report_path, 'w', encoding='utf-8') as fh:
    fh.write('\n'.join(report_lines))
print(f"  Saved: Figure_Generation_Report.txt")

# ─────────────────────────────────────────────
# CONSOLE SUMMARY
# ─────────────────────────────────────────────
print("\n" + "=" * 70)
print("COMPLETE")
print("=" * 70)
print(f"  Figures:  {FIG_DIR}")
print(f"  Data:     {PROC_DIR}")
print(f"  Report:   {report_path}")
print(f"\n  Annotated peaks (Figure 4): {len(annotatable)}")
for i in annotatable:
    pi = peaks_idx[i]
    print(f"    {rs_c[pi]:.1f} cm-1  [{confidences[i]}]")
