"""
Investigation Script:
1. Refined peak table (publication-quality)
2. Zoom plot of 650-800 cm-1 region
3. Investigate anomalous 90%-60s SNR
"""
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from scipy.signal import find_peaks, savgol_filter
from scipy import sparse
from scipy.sparse.linalg import spsolve
import os

BASE_DIR = r"c:\Users\sukes\Downloads\nam-new"
PROCESSED_DIR = os.path.join(BASE_DIR, "Analysis", "Processed")
FIGURES_DIR = os.path.join(BASE_DIR, "Analysis", "Figures")
os.chdir(BASE_DIR)

plt.rcParams.update({'font.size': 11, 'axes.linewidth': 1.2, 'figure.dpi': 150,
                     'savefig.dpi': 300, 'savefig.bbox': 'tight'})

# --- Helper functions ---
def extract_spectrum(xlsx_path):
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    rs, inten = [], []
    for row in ws.iter_rows(min_row=100, values_only=True):
        try:
            rs.append(float(row[3]))
            inten.append(float(row[7]))
        except (TypeError, ValueError, IndexError):
            continue
    wb.close()
    return np.array(rs), np.array(inten)

def baseline_als(y, lam=1e5, p=0.001, niter=10):
    L = len(y)
    D = sparse.diags([1, -2, 1], [0, -1, -2], shape=(L, L - 2))
    w = np.ones(L)
    for i in range(niter):
        W = sparse.spdiags(w, 0, L, L)
        Z = W + lam * D.dot(D.T)
        z = spsolve(Z, w * y)
        w = p * (y > z) + (1 - p) * (y < z)
    return z

def process_spectrum(rs, intensity, region=(400, 1800)):
    # Crop
    mask = (rs >= region[0]) & (rs <= region[1])
    rs_c, int_c = rs[mask], intensity[mask]
    # Baseline
    bl = baseline_als(int_c)
    int_bc = int_c - bl
    # Smooth
    int_sm = savgol_filter(int_bc, window_length=11, polyorder=3)
    # Normalize
    imin, imax = int_sm.min(), int_sm.max()
    int_norm = (int_sm - imin) / (imax - imin) if imax > imin else np.zeros_like(int_sm)
    return rs_c, int_norm, int_c, bl

# ============================================================================
# PART 1: Refined Peak Table
# ============================================================================
print("=" * 80)
print("PART 1: REFINED PEAK TABLE (Publication Quality)")
print("=" * 80)

# Load and process NAM average
nam_dir = r"70p-60s-5ac-diifrentpoint"
nam_files = sorted([f for f in os.listdir(nam_dir) if f.endswith('.xlsx')])

processed_spectra = []
for f in nam_files:
    rs, intensity = extract_spectrum(os.path.join(nam_dir, f))
    rs_c, int_norm, _, _ = process_spectrum(rs, intensity)
    processed_spectra.append(int_norm)

rs_common = rs_c
nam_mean = np.mean(processed_spectra, axis=0)
nam_std = np.std(processed_spectra, axis=0)

# Table A: All peaks (current - prominence=0.02, distance=10)
peaks_all, props_all = find_peaks(nam_mean, prominence=0.02, distance=10)

# Table B: Publication peaks (prominence=0.05, distance=15)
peaks_pub, props_pub = find_peaks(nam_mean, prominence=0.05, distance=15)

# Table C: Conservative (prominence=0.08, distance=20)
peaks_con, props_con = find_peaks(nam_mean, prominence=0.08, distance=20)

print(f"\n  Table A (all peaks):         {len(peaks_all)} peaks (prom>0.02, dist>10)")
print(f"  Table B (publication):       {len(peaks_pub)} peaks (prom>0.05, dist>15)")
print(f"  Table C (conservative):      {len(peaks_con)} peaks (prom>0.08, dist>20)")

print(f"\n  TABLE B - PUBLICATION PEAKS:")
print(f"  {'#':<4} {'Position (cm-1)':<18} {'Intensity':<12} {'Prominence':<12}")
print("  " + "-" * 46)
for i, (p, prom) in enumerate(zip(peaks_pub, props_pub['prominences']), 1):
    print(f"  {i:<4} {rs_common[p]:<18.1f} {nam_mean[p]:<12.4f} {prom:<12.4f}")

# Save publication peak table
df_pub = pd.DataFrame({
    'Peak #': range(1, len(peaks_pub) + 1),
    'Position (cm-1)': rs_common[peaks_pub],
    'Normalized Intensity': nam_mean[peaks_pub],
    'Prominence': props_pub['prominences']
})
df_pub.to_excel(os.path.join(PROCESSED_DIR, 'NAM_peak_list_publication.xlsx'), index=False)
print(f"\n  Saved: NAM_peak_list_publication.xlsx")

# ============================================================================
# PART 2: Zoom Plot 650-800 cm-1 (Cell Wall Region)
# ============================================================================
print("\n" + "=" * 80)
print("PART 2: ZOOM PLOT 650-800 cm-1 (Bacterial Cell Wall Region)")
print("=" * 80)

# Get raw + processed data for zoom
zoom_min, zoom_max = 620, 820

fig, axes = plt.subplots(3, 1, figsize=(8, 8), sharex=True)
fig.subplots_adjust(hspace=0.12)

# Panel (a): Raw spectra (all 3 spots)
colors = ['#e41a1c', '#377eb8', '#4daf4a']
for i, f in enumerate(nam_files):
    rs, intensity = extract_spectrum(os.path.join(nam_dir, f))
    mask = (rs >= zoom_min) & (rs <= zoom_max)
    axes[0].plot(rs[mask], intensity[mask], color=colors[i], lw=0.8,
                 label=f'Spot {i+1}', alpha=0.8)

axes[0].set_ylabel('Raw Counts')
axes[0].set_title('(a) Raw Spectra — 620-820 cm$^{-1}$ Region', fontweight='bold', loc='left')
axes[0].legend(loc='upper right', fontsize=9)

# Panel (b): Processed individual spectra
for i, int_norm in enumerate(processed_spectra):
    mask = (rs_common >= zoom_min) & (rs_common <= zoom_max)
    axes[1].plot(rs_common[mask], int_norm[mask], color=colors[i], lw=0.8,
                 label=f'Spot {i+1}', alpha=0.8)

axes[1].set_ylabel('Normalized Int.')
axes[1].set_title('(b) Processed Spectra (baseline corrected, smoothed, normalized)',
                  fontweight='bold', loc='left')
axes[1].legend(loc='upper right', fontsize=9)

# Panel (c): Mean spectrum with std
mask = (rs_common >= zoom_min) & (rs_common <= zoom_max)
rs_zoom = rs_common[mask]
mean_zoom = nam_mean[mask]
std_zoom = nam_std[mask]

axes[2].fill_between(rs_zoom, mean_zoom - std_zoom, mean_zoom + std_zoom,
                     alpha=0.3, color='black')
axes[2].plot(rs_zoom, mean_zoom, 'k-', lw=1.2)

# Detect peaks in this region
peaks_zoom, props_zoom = find_peaks(mean_zoom, prominence=0.02, distance=8)
for p in peaks_zoom:
    axes[2].axvline(rs_zoom[p], color='red', lw=0.5, ls='--', alpha=0.5)
    axes[2].annotate(f'{rs_zoom[p]:.0f}', xy=(rs_zoom[p], mean_zoom[p]),
                     xytext=(0, 8), textcoords='offset points',
                     ha='center', fontsize=9, fontweight='bold', color='red')
    axes[2].plot(rs_zoom[p], mean_zoom[p], 'rv', markersize=5)

axes[2].set_ylabel('Normalized Int.')
axes[2].set_xlabel('Raman Shift (cm$^{-1}$)')
axes[2].set_title('(c) Mean Spectrum with Peak Positions', fontweight='bold', loc='left')
axes[2].set_xlim(zoom_min, zoom_max)

# Add annotation about cell wall band
axes[2].axvspan(725, 735, alpha=0.15, color='green')
axes[2].text(730, axes[2].get_ylim()[1] * 0.9, 'Cell wall\n(literature)',
             ha='center', fontsize=8, color='green', fontstyle='italic')

plt.tight_layout()
plt.savefig(os.path.join(FIGURES_DIR, 'Figure_Zoom_650_800_CellWall.png'), dpi=300, facecolor='white')
plt.savefig(os.path.join(FIGURES_DIR, 'Figure_Zoom_650_800_CellWall.pdf'), facecolor='white')
plt.close()

print(f"\n  Peaks in 620-820 cm-1 region:")
for p in peaks_zoom:
    print(f"    {rs_zoom[p]:.1f} cm-1 (intensity: {mean_zoom[p]:.3f})")

print(f"\n  Literature cell wall band: 725-735 cm-1")
# Check if any peak falls in 725-735
cell_wall_peaks = [rs_zoom[p] for p in peaks_zoom if 720 <= rs_zoom[p] <= 740]
if cell_wall_peaks:
    print(f"  >>> PEAK DETECTED in cell wall region: {cell_wall_peaks}")
else:
    print(f"  >>> No peak at 725-735 cm-1. Nearest peak: {rs_zoom[peaks_zoom[np.argmin(np.abs(rs_zoom[peaks_zoom] - 730))]]:.1f} cm-1")

print(f"\n  [OK] Figure saved: Figure_Zoom_650_800_CellWall.png")

# ============================================================================
# PART 3: Investigate 90%-60s Anomalous SNR
# ============================================================================
print("\n" + "=" * 80)
print("PART 3: INVESTIGATING 90%-60s ANOMALOUS SNR")
print("=" * 80)

# Load 90%-60s data
anom_dir = r"90p-60s-5a-samespot"
anom_files = sorted([f for f in os.listdir(anom_dir) if f.endswith('.xlsx')])
print(f"\n  Folder: {anom_dir}")
print(f"  Files: {len(anom_files)}")

# Also load 70%-60s (good data) for comparison
good_dir = r"70p-60s-5ac-diifrentpoint"
good_files = sorted([f for f in os.listdir(good_dir) if f.endswith('.xlsx')])

fig, axes = plt.subplots(3, 2, figsize=(14, 10))
fig.suptitle('Investigation: 90% 60s (Anomalous) vs 70% 60s (Good)', fontweight='bold', fontsize=13)

# Load raw spectra
anom_spectra_raw = []
for f in anom_files:
    rs, intensity = extract_spectrum(os.path.join(anom_dir, f))
    anom_spectra_raw.append((rs, intensity, f))

good_spectra_raw = []
for f in good_files:
    rs, intensity = extract_spectrum(os.path.join(good_dir, f))
    good_spectra_raw.append((rs, intensity, f))

# Row 1: Full raw spectra
for i, (rs, inten, f) in enumerate(anom_spectra_raw):
    mask = (rs >= 200) & (rs <= 2000)
    axes[0, 0].plot(rs[mask], inten[mask], lw=0.6, alpha=0.7, label=f'Spot {i+1}')
axes[0, 0].set_title('90%, 60s — Raw Spectra', fontweight='bold')
axes[0, 0].set_ylabel('Counts')
axes[0, 0].legend(fontsize=7)

for i, (rs, inten, f) in enumerate(good_spectra_raw):
    mask = (rs >= 200) & (rs <= 2000)
    axes[0, 1].plot(rs[mask], inten[mask], lw=0.6, alpha=0.7, label=f'Spot {i+1}')
axes[0, 1].set_title('70%, 60s — Raw Spectra', fontweight='bold')
axes[0, 1].set_ylabel('Counts')
axes[0, 1].legend(fontsize=7)

# Row 2: Signal region (900-1100) vs Noise region (1700-1800)
for i, (rs, inten, f) in enumerate(anom_spectra_raw):
    sig_mask = (rs >= 900) & (rs <= 1100)
    axes[1, 0].plot(rs[sig_mask], inten[sig_mask], lw=0.7, alpha=0.7)
axes[1, 0].set_title('90%, 60s — Signal Region (900-1100)', fontweight='bold')
axes[1, 0].set_ylabel('Counts')
axes[1, 0].axhline(np.mean([inten[(rs >= 900) & (rs <= 1100)].max() for rs, inten, _ in anom_spectra_raw]),
                   color='red', ls='--', lw=0.8, label='Mean max')
axes[1, 0].legend(fontsize=8)

for i, (rs, inten, f) in enumerate(good_spectra_raw):
    sig_mask = (rs >= 900) & (rs <= 1100)
    axes[1, 1].plot(rs[sig_mask], inten[sig_mask], lw=0.7, alpha=0.7)
axes[1, 1].set_title('70%, 60s — Signal Region (900-1100)', fontweight='bold')
axes[1, 1].set_ylabel('Counts')

# Row 3: Noise region
for i, (rs, inten, f) in enumerate(anom_spectra_raw):
    noise_mask = (rs >= 1700) & (rs <= 1800)
    axes[2, 0].plot(rs[noise_mask], inten[noise_mask], lw=0.7, alpha=0.7)
axes[2, 0].set_title('90%, 60s — Noise Region (1700-1800)', fontweight='bold')
axes[2, 0].set_xlabel('Raman Shift (cm$^{-1}$)')
axes[2, 0].set_ylabel('Counts')

for i, (rs, inten, f) in enumerate(good_spectra_raw):
    noise_mask = (rs >= 1700) & (rs <= 1800)
    axes[2, 1].plot(rs[noise_mask], inten[noise_mask], lw=0.7, alpha=0.7)
axes[2, 1].set_title('70%, 60s — Noise Region (1700-1800)', fontweight='bold')
axes[2, 1].set_xlabel('Raman Shift (cm$^{-1}$)')
axes[2, 1].set_ylabel('Counts')

plt.tight_layout()
plt.savefig(os.path.join(FIGURES_DIR, 'Figure_Investigation_90p60s.png'), dpi=300, facecolor='white')
plt.savefig(os.path.join(FIGURES_DIR, 'Figure_Investigation_90p60s.pdf'), facecolor='white')
plt.close()

# Print diagnostics
print("\n  DIAGNOSTIC COMPARISON:")
print(f"\n  {'Metric':<30} {'90%, 60s':<15} {'70%, 60s':<15}")
print("  " + "-" * 60)

# Max intensity
anom_max = [inten.max() for _, inten, _ in anom_spectra_raw]
good_max = [inten.max() for _, inten, _ in good_spectra_raw]
print(f"  {'Max raw counts':<30} {np.mean(anom_max):<15.0f} {np.mean(good_max):<15.0f}")

# Check for saturation (65535 = 16-bit max)
anom_sat = [np.sum(inten >= 65535) for _, inten, _ in anom_spectra_raw]
good_sat = [np.sum(inten >= 65535) for _, inten, _ in good_spectra_raw]
print(f"  {'Saturated pixels (=65535)':<30} {np.mean(anom_sat):<15.1f} {np.mean(good_sat):<15.1f}")

# Signal in 900-1100
anom_signal = []
anom_noise_std = []
for rs, inten, _ in anom_spectra_raw:
    sig_mask = (rs >= 900) & (rs <= 1100)
    noise_mask = (rs >= 1700) & (rs <= 1800)
    sig = np.max(inten[sig_mask]) - np.median(inten[noise_mask])
    noise = np.std(inten[noise_mask])
    anom_signal.append(sig)
    anom_noise_std.append(noise)

good_signal = []
good_noise_std = []
for rs, inten, _ in good_spectra_raw:
    sig_mask = (rs >= 900) & (rs <= 1100)
    noise_mask = (rs >= 1700) & (rs <= 1800)
    sig = np.max(inten[sig_mask]) - np.median(inten[noise_mask])
    noise = np.std(inten[noise_mask])
    good_signal.append(sig)
    good_noise_std.append(noise)

print(f"  {'Signal (900-1100 max-median)':<30} {np.mean(anom_signal):<15.0f} {np.mean(good_signal):<15.0f}")
print(f"  {'Noise std (1700-1800)':<30} {np.mean(anom_noise_std):<15.1f} {np.mean(good_noise_std):<15.1f}")
print(f"  {'SNR (signal/noise)':<30} {np.mean(anom_signal)/np.mean(anom_noise_std):<15.1f} {np.mean(good_signal)/np.mean(good_noise_std):<15.1f}")

# Median intensity in noise region
anom_noise_med = [np.median(inten[(rs >= 1700) & (rs <= 1800)]) for rs, inten, _ in anom_spectra_raw]
good_noise_med = [np.median(inten[(rs >= 1700) & (rs <= 1800)]) for rs, inten, _ in good_spectra_raw]
print(f"  {'Noise median (1700-1800)':<30} {np.mean(anom_noise_med):<15.0f} {np.mean(good_noise_med):<15.0f}")

# Check if signal is BELOW noise median (explains negative SNR)
print(f"\n  DIAGNOSIS:")
if np.mean(anom_signal) < 0:
    print(f"  >>> NEGATIVE SIGNAL: Signal peak (900-1100) is BELOW noise median")
    print(f"      This means the spectrum is dominated by fluorescence/background")
    print(f"      The Raman peaks at 900-1100 cm-1 are buried under the baseline")
    print(f"      CONCLUSION: 90% power causes excessive fluorescence/burning")
elif np.mean(anom_noise_std) > np.mean(good_noise_std) * 3:
    print(f"  >>> HIGH NOISE: Noise region has unusually high variance")
    print(f"      CONCLUSION: Possible detector saturation or sample damage")
else:
    print(f"  >>> LOW SIGNAL relative to noise floor")
    print(f"      The 900-1100 region peaks are not prominent at this condition")
    print(f"      CONCLUSION: Different spectral shape or sample degradation at 90% power")

print(f"\n  [OK] Figure saved: Figure_Investigation_90p60s.png")

# ============================================================================
# FINAL SUMMARY
# ============================================================================
print("\n" + "=" * 80)
print("INVESTIGATION COMPLETE")
print("=" * 80)

print(f"\n  1. PEAK TABLES:")
print(f"     Table A (all):          {len(peaks_all)} peaks")
print(f"     Table B (publication):  {len(peaks_pub)} peaks  <-- USE THIS")
print(f"     Table C (conservative): {len(peaks_con)} peaks")

print(f"\n  2. CELL WALL REGION (650-800 cm-1):")
print(f"     Peaks found in 620-820 region: {len(peaks_zoom)}")
for p in peaks_zoom:
    marker = " *** CELL WALL BAND" if 720 <= rs_zoom[p] <= 740 else ""
    print(f"       {rs_zoom[p]:.1f} cm-1{marker}")

print(f"\n  3. 90%-60s ANOMALY:")
print(f"     Issue identified and diagnostic figure generated")
print(f"     See Figure_Investigation_90p60s.png for details")

print(f"\n  Files generated:")
print(f"    - NAM_peak_list_publication.xlsx")
print(f"    - Figure_Zoom_650_800_CellWall.png/pdf")
print(f"    - Figure_Investigation_90p60s.png/pdf")
