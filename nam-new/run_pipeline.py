"""Run the complete Raman processing pipeline."""
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from scipy.signal import find_peaks, savgol_filter
from scipy import sparse
from scipy.sparse.linalg import spsolve
import os

BASE_DIR = r"c:\Users\sukes\Downloads\nam-new"
PROCESSED_DIR = os.path.join(BASE_DIR, "Analysis", "Processed")
FIGURES_DIR = os.path.join(BASE_DIR, "Analysis", "Figures")
os.makedirs(PROCESSED_DIR, exist_ok=True)
os.chdir(BASE_DIR)

plt.rcParams.update({'font.size': 11, 'axes.linewidth': 1.2, 'figure.dpi': 150,
                     'savefig.dpi': 300, 'savefig.bbox': 'tight'})

# ===== STEP 1: Extract =====
print("="*80)
print("STEP 1: Extracting Raman Shift + Dark Subtracted")
print("="*80)

def extract_spectrum(xlsx_path):
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    rs, inten = [], []
    for row in ws.iter_rows(min_row=100, values_only=True):
        try:
            rs.append(float(row[3]))
            inten.append(float(row[7]))
        except (TypeError, ValueError, IndexError):
            continue
    wb.close()
    return np.array(rs), np.array(inten)

nam_dir = r"70p-60s-5ac-diifrentpoint"
nam_files = sorted([f for f in os.listdir(nam_dir) if f.endswith('.xlsx')])
nam_raw = {}
for f in nam_files:
    rs, intensity = extract_spectrum(os.path.join(nam_dir, f))
    nam_raw[f] = {'raman_shift': rs, 'intensity': intensity}
    print(f"  NAM: {f} -> {len(rs)} pts")

glass_dir = r"glass slide empty\empty slide"
glass_files = sorted([f for f in os.listdir(glass_dir) if f.endswith('.xlsx')])
glass_raw = {}
for f in glass_files:
    rs, intensity = extract_spectrum(os.path.join(glass_dir, f))
    glass_raw[f] = {'raman_shift': rs, 'intensity': intensity}
print(f"  Glass: {len(glass_raw)} spectra loaded")

# ===== STEPS 2 & 3: Clean + Crop =====
print("\n" + "="*80)
print("STEPS 2-3: Cleaning + Cropping to 400-1800 cm-1")
print("="*80)

def clean_and_crop(raman_shift, intensity, region=(400, 1800)):
    valid = ~(np.isnan(raman_shift) | np.isnan(intensity))
    rs, inten = raman_shift[valid], intensity[valid]
    mask = (rs >= region[0]) & (rs <= region[1])
    return rs[mask], inten[mask]

nam_fp = {}
for f, d in nam_raw.items():
    rs, inten = clean_and_crop(d['raman_shift'], d['intensity'])
    nam_fp[f] = {'raman_shift': rs, 'intensity': inten}
    print(f"  {f}: {len(rs)} pts in 400-1800 cm-1")

glass_fp = {}
for f, d in glass_raw.items():
    rs, inten = clean_and_crop(d['raman_shift'], d['intensity'])
    glass_fp[f] = {'raman_shift': rs, 'intensity': inten}

# ===== STEP 4: ALS Baseline =====
print("\n" + "="*80)
print("STEP 4: ALS Baseline Correction (lam=1e5, p=0.001, iter=10)")
print("="*80)

def baseline_als(y, lam=1e5, p=0.001, niter=10):
    L = len(y)
    D = sparse.diags([1, -2, 1], [0, -1, -2], shape=(L, L - 2))
    w = np.ones(L)
    for i in range(niter):
        W = sparse.spdiags(w, 0, L, L)
        Z = W + lam * D.dot(D.T)
        z = spsolve(Z, w * y)
        w = p * (y > z) + (1 - p) * (y < z)
    return z

nam_bc = {}
for f, d in nam_fp.items():
    bl = baseline_als(d['intensity'])
    nam_bc[f] = {'raman_shift': d['raman_shift'], 'intensity': d['intensity'] - bl,
                 'baseline': bl, 'raw': d['intensity']}
    print(f"  {f}: baseline subtracted")

glass_bc = {}
for f, d in glass_fp.items():
    bl = baseline_als(d['intensity'])
    glass_bc[f] = {'raman_shift': d['raman_shift'], 'intensity': d['intensity'] - bl}

# ===== STEP 5: Smoothing =====
print("\n" + "="*80)
print("STEP 5: Savitzky-Golay Smoothing (window=11, poly=3)")
print("="*80)

nam_sm = {}
for f, d in nam_bc.items():
    sm = savgol_filter(d['intensity'], window_length=11, polyorder=3)
    nam_sm[f] = {'raman_shift': d['raman_shift'], 'intensity': sm, 'unsmoothed': d['intensity']}

glass_sm = {}
for f, d in glass_bc.items():
    sm = savgol_filter(d['intensity'], window_length=11, polyorder=3)
    glass_sm[f] = {'raman_shift': d['raman_shift'], 'intensity': sm}

print("  All spectra smoothed.")

# ===== STEP 6: Normalization =====
print("\n" + "="*80)
print("STEP 6: Min-Max Normalization")
print("="*80)

def normalize_minmax(intensity):
    imin, imax = intensity.min(), intensity.max()
    if imax - imin == 0:
        return np.zeros_like(intensity)
    return (intensity - imin) / (imax - imin)

nam_norm = {}
for f, d in nam_sm.items():
    nam_norm[f] = {'raman_shift': d['raman_shift'], 'intensity': normalize_minmax(d['intensity'])}

glass_norm = {}
for f, d in glass_sm.items():
    glass_norm[f] = {'raman_shift': d['raman_shift'], 'intensity': normalize_minmax(d['intensity'])}

print("  All spectra normalized to [0, 1].")

# ===== STEP 7: Average =====
print("\n" + "="*80)
print("STEP 7: Averaging Replicates")
print("="*80)

rs_common = list(nam_norm.values())[0]['raman_shift']
nam_stack = np.array([d['intensity'] for d in nam_norm.values()])
nam_mean = nam_stack.mean(axis=0)
nam_std = nam_stack.std(axis=0)
print(f"  NAM: {nam_stack.shape[0]} spectra averaged")

rs_glass = list(glass_norm.values())[0]['raman_shift']
glass_stack = np.array([d['intensity'] for d in glass_norm.values()])
glass_mean = glass_stack.mean(axis=0)
glass_std = glass_stack.std(axis=0)
print(f"  Glass: {glass_stack.shape[0]} spectra averaged")

# ===== STEP 8: Peak Detection =====
print("\n" + "="*80)
print("STEP 8: Peak Detection (prominence=0.02, distance=10)")
print("="*80)

peaks, props = find_peaks(nam_mean, prominence=0.02, distance=10)
peak_pos = rs_common[peaks]
peak_int = nam_mean[peaks]
peak_prom = props['prominences']

print(f"  Peaks detected: {len(peaks)}")
print(f"  Positions: {[f'{p:.0f}' for p in peak_pos]}")

# ===== FIGURES =====
print("\n" + "="*80)
print("GENERATING FIGURES")
print("="*80)

# Figure 1: Glass vs NAM
fig, ax = plt.subplots(figsize=(10, 5))
ax.fill_between(rs_glass, glass_mean - glass_std, glass_mean + glass_std, alpha=0.3, color='blue')
ax.plot(rs_glass, glass_mean, 'b-', lw=1.0, label='Glass (mean)')
ax.fill_between(rs_common, nam_mean - nam_std, nam_mean + nam_std, alpha=0.3, color='red')
ax.plot(rs_common, nam_mean, 'r-', lw=1.0, label='NAM (mean)')
ax.plot(peak_pos, peak_int, 'kv', markersize=5)
for pos, inten in zip(peak_pos, peak_int):
    if inten > 0.1:
        ax.annotate(f'{pos:.0f}', xy=(pos, inten), xytext=(0, 8),
                    textcoords='offset points', ha='center', fontsize=7, color='darkred')
ax.set_xlabel('Raman Shift (cm$^{-1}$)')
ax.set_ylabel('Normalized Intensity')
ax.set_title('Figure 1: Glass vs NAM (Processed)', fontweight='bold')
ax.legend(loc='upper right')
ax.set_xlim(400, 1800)
ax.set_ylim(-0.05, 1.1)
plt.tight_layout()
plt.savefig(os.path.join(FIGURES_DIR, 'Figure1_Glass_vs_NAM.png'), dpi=300, facecolor='white')
plt.savefig(os.path.join(FIGURES_DIR, 'Figure1_Glass_vs_NAM.pdf'), facecolor='white')
plt.close()
print("  [OK] Figure1_Glass_vs_NAM")

# Figure 4: Final Raman
fig, ax = plt.subplots(figsize=(10, 4.5))
ax.fill_between(rs_common, nam_mean - nam_std, nam_mean + nam_std, alpha=0.2, color='black')
ax.plot(rs_common, nam_mean, 'k-', lw=1.0)
for pos, inten, prom in zip(peak_pos, peak_int, peak_prom):
    if prom > 0.03:
        ax.annotate(f'{pos:.0f}', xy=(pos, inten), xytext=(0, 8),
                    textcoords='offset points', ha='center', fontsize=8, fontweight='bold', color='red')
        ax.plot(pos, inten, 'rv', markersize=4)
ax.set_xlabel('Raman Shift (cm$^{-1}$)', fontweight='bold')
ax.set_ylabel('Normalized Intensity (a.u.)', fontweight='bold')
ax.set_xlim(400, 1800)
ax.set_ylim(-0.02, 1.05)
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)
plt.tight_layout()
plt.savefig(os.path.join(FIGURES_DIR, 'Figure4_Final_Raman.png'), dpi=300, facecolor='white')
plt.savefig(os.path.join(FIGURES_DIR, 'Figure4_Final_Raman.pdf'), facecolor='white')
plt.close()
print("  [OK] Figure4_Final_Raman")

# Processing Pipeline Figure
f0 = list(nam_raw.keys())[0]
fig, axes = plt.subplots(5, 1, figsize=(10, 12), sharex=True)
fig.subplots_adjust(hspace=0.15)

rs_r = nam_raw[f0]['raman_shift']
int_r = nam_raw[f0]['intensity']
m_r = (rs_r >= 400) & (rs_r <= 1800)
axes[0].plot(rs_r[m_r], int_r[m_r], 'b-', lw=0.6)
axes[0].set_title('(a) Raw Spectrum', fontweight='bold', loc='left')
axes[0].set_ylabel('Counts')

d_bc = nam_bc[f0]
axes[1].plot(d_bc['raman_shift'], d_bc['raw'], 'b-', lw=0.5, alpha=0.5, label='Raw')
axes[1].plot(d_bc['raman_shift'], d_bc['baseline'], 'r-', lw=1.5, label='Baseline')
axes[1].set_title('(b) ALS Baseline Fitting', fontweight='bold', loc='left')
axes[1].set_ylabel('Counts')
axes[1].legend(fontsize=8)

axes[2].plot(d_bc['raman_shift'], d_bc['intensity'], 'g-', lw=0.6)
axes[2].axhline(0, color='gray', lw=0.5, ls='--')
axes[2].set_title('(c) Baseline Corrected', fontweight='bold', loc='left')
axes[2].set_ylabel('Intensity')

d_sm = nam_sm[f0]
axes[3].plot(d_sm['raman_shift'], d_sm['unsmoothed'], 'gray', lw=0.4, alpha=0.5, label='Noisy')
axes[3].plot(d_sm['raman_shift'], d_sm['intensity'], 'k-', lw=0.8, label='Smoothed')
axes[3].set_title('(d) Savitzky-Golay Smoothing', fontweight='bold', loc='left')
axes[3].set_ylabel('Intensity')
axes[3].legend(fontsize=8)

d_nm = nam_norm[f0]
axes[4].plot(d_nm['raman_shift'], d_nm['intensity'], 'k-', lw=0.8)
axes[4].set_title('(e) Min-Max Normalized', fontweight='bold', loc='left')
axes[4].set_xlabel('Raman Shift (cm$^{-1}$)')
axes[4].set_ylabel('Norm. Int.')
axes[4].set_xlim(400, 1800)

plt.savefig(os.path.join(FIGURES_DIR, 'Figure_Processing_Pipeline.png'), dpi=300, facecolor='white')
plt.savefig(os.path.join(FIGURES_DIR, 'Figure_Processing_Pipeline.pdf'), facecolor='white')
plt.close()
print("  [OK] Figure_Processing_Pipeline")

# ===== SAVE DATA =====
print("\n" + "="*80)
print("SAVING PROCESSED DATA")
print("="*80)

for i, (f, d) in enumerate(nam_norm.items(), 1):
    df = pd.DataFrame({'Raman Shift (cm-1)': d['raman_shift'], 'Normalized Intensity': d['intensity']})
    out = f"NAM_70P_60S_5AC_spot{i}_processed.xlsx"
    df.to_excel(os.path.join(PROCESSED_DIR, out), index=False)
    print(f"  {out}")

df_avg = pd.DataFrame({'Raman Shift (cm-1)': rs_common, 'Mean Intensity': nam_mean, 'Std': nam_std})
df_avg.to_excel(os.path.join(PROCESSED_DIR, 'NAM_average.xlsx'), index=False)
print("  NAM_average.xlsx")

df_gl = pd.DataFrame({'Raman Shift (cm-1)': rs_glass, 'Mean Intensity': glass_mean, 'Std': glass_std})
df_gl.to_excel(os.path.join(PROCESSED_DIR, 'glass_processed.xlsx'), index=False)
print("  glass_processed.xlsx")

df_pk = pd.DataFrame({'Peak Position (cm-1)': peak_pos, 'Intensity': peak_int, 'Prominence': peak_prom})
df_pk.to_excel(os.path.join(PROCESSED_DIR, 'NAM_peak_list.xlsx'), index=False)
print("  NAM_peak_list.xlsx")

# ===== FINAL SUMMARY =====
print("\n" + "="*80)
print("PIPELINE COMPLETE")
print("="*80)
print(f"\nProcessed Files: {PROCESSED_DIR}")
for f in sorted(os.listdir(PROCESSED_DIR)):
    print(f"  {f}")
print(f"\nFigures: {FIGURES_DIR}")
for f in sorted(os.listdir(FIGURES_DIR)):
    if 'Figure' in f and f.endswith('.png'):
        print(f"  {f}")
print(f"\nPeak List:")
for pos, prom in zip(peak_pos, peak_prom):
    if prom > 0.03:
        print(f"  {pos:.0f} cm-1 (prominence: {prom:.3f})")
