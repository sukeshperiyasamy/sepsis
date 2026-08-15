import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from scipy.signal import find_peaks
import os

os.chdir(r"c:\Users\sukes\Downloads\nam-new")

plt.rcParams.update({
    'font.size': 11,
    'axes.linewidth': 1.2,
    'xtick.major.width': 1.2,
    'ytick.major.width': 1.2,
    'figure.dpi': 150,
    'savefig.dpi': 300,
    'savefig.bbox': 'tight'
})

def extract_spectrum(xlsx_path):
    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    raman_shift = []
    intensity = []
    for row in ws.iter_rows(min_row=100, values_only=True):
        try:
            rs = float(row[3])
            ds = float(row[7])
            raman_shift.append(rs)
            intensity.append(ds)
        except (TypeError, ValueError, IndexError):
            continue
    wb.close()
    return np.array(raman_shift), np.array(intensity)

def load_folder_spectra(folder_path):
    spectra = []
    files = sorted([f for f in os.listdir(folder_path) if f.endswith('.xlsx')])
    for f in files:
        rs, intensity = extract_spectrum(os.path.join(folder_path, f))
        spectra.append({'file': f, 'raman_shift': rs, 'intensity': intensity})
    print(f"  Loaded {len(spectra)} from: {folder_path}")
    return spectra

xmin, xmax = 200, 1800

print("=" * 80)
print("GENERATING ALL FIGURES")
print("=" * 80)

# ===== FIGURE S1: Glass vs NAM =====
print("\n[1/5] Figure S1: Glass vs NAM...")
glass_spectra = load_folder_spectra(r"glass slide empty\empty slide")
nam_spectra = load_folder_spectra(r"70p-60s-5ac-diifrentpoint")

fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 7), sharex=True)
fig.subplots_adjust(hspace=0.08)

glass_colors = plt.cm.Blues(np.linspace(0.4, 0.9, 5))
for i, sp in enumerate(glass_spectra):
    mask = (sp['raman_shift'] >= xmin) & (sp['raman_shift'] <= xmax)
    ax1.plot(sp['raman_shift'][mask], sp['intensity'][mask],
             linewidth=0.7, alpha=0.8, color=glass_colors[i], label=f"Glass {i+1}")

ax1.set_ylabel('Intensity (a.u.)', fontsize=12)
ax1.set_title('(a) Empty Glass Slide', fontsize=12, fontweight='bold', loc='left')
ax1.legend(loc='upper right', fontsize=9, framealpha=0.9)
ax1.set_xlim(xmin, xmax)

nam_colors = ['#e41a1c', '#377eb8', '#4daf4a']
for i, sp in enumerate(nam_spectra):
    mask = (sp['raman_shift'] >= xmin) & (sp['raman_shift'] <= xmax)
    ax2.plot(sp['raman_shift'][mask], sp['intensity'][mask],
             linewidth=0.7, alpha=0.85, color=nam_colors[i], label=f"NAM Spot {i+1}")

ax2.set_xlabel('Raman Shift (cm$^{-1}$)', fontsize=12)
ax2.set_ylabel('Intensity (a.u.)', fontsize=12)
ax2.set_title('(b) NAM Sample (70% power, 60s, 5 acc)', fontsize=12, fontweight='bold', loc='left')
ax2.legend(loc='upper right', fontsize=9, framealpha=0.9)

plt.savefig(r'Analysis\Figures\Figure_S1_Glass_vs_NAM.png', dpi=300, facecolor='white')
plt.savefig(r'Analysis\Figures\Figure_S1_Glass_vs_NAM.pdf', facecolor='white')
plt.close()
print("  [OK] Saved.")

# ===== FIGURE 2: Reproducibility =====
print("\n[2/5] Figure 2: Reproducibility...")
same_spot = load_folder_spectra(r"90p-30s-5acc-samespot")
diff_spot = load_folder_spectra(r"90p-30s-5ac-diifrentpoint")

fig, (ax1, ax2) = plt.subplots(2, 1, figsize=(10, 7), sharex=True)
fig.subplots_adjust(hspace=0.08)

same_colors = plt.cm.Reds(np.linspace(0.3, 0.9, len(same_spot)))
for i, sp in enumerate(same_spot):
    mask = (sp['raman_shift'] >= xmin) & (sp['raman_shift'] <= xmax)
    ax1.plot(sp['raman_shift'][mask], sp['intensity'][mask],
             linewidth=0.5, alpha=0.7, color=same_colors[i])

ax1.set_ylabel('Intensity (a.u.)', fontsize=12)
ax1.set_title(f'(a) Same Spot - {len(same_spot)} repeated measurements (90%, 30s, 5acc)',
              fontsize=11, fontweight='bold', loc='left')
ax1.set_xlim(xmin, xmax)

diff_colors = plt.cm.Set1(np.linspace(0, 1, len(diff_spot)))
for i, sp in enumerate(diff_spot):
    mask = (sp['raman_shift'] >= xmin) & (sp['raman_shift'] <= xmax)
    ax2.plot(sp['raman_shift'][mask], sp['intensity'][mask],
             linewidth=0.7, alpha=0.8, color=diff_colors[i], label=f"Spot {i+1}")

ax2.set_xlabel('Raman Shift (cm$^{-1}$)', fontsize=12)
ax2.set_ylabel('Intensity (a.u.)', fontsize=12)
ax2.set_title(f'(b) Different Spots - {len(diff_spot)} locations (90%, 30s, 5acc)',
              fontsize=11, fontweight='bold', loc='left')
ax2.legend(loc='upper right', fontsize=9, framealpha=0.9)

plt.savefig(r'Analysis\Figures\Figure_2_Reproducibility.png', dpi=300, facecolor='white')
plt.savefig(r'Analysis\Figures\Figure_2_Reproducibility.pdf', facecolor='white')
plt.close()
print("  [OK] Saved.")

# ===== FIGURE 3: Conditions Comparison =====
print("\n[3/5] Figure 3: Mean Spectra per Condition...")
conditions = {
    '70%, 10s, 5acc': load_folder_spectra(r"NAM-70p-10s-5ac"),
    '70%, 10s, 3acc': load_folder_spectra(r"NAM-70p-10s-3acc"),
    '90%, 10s, 5acc': load_folder_spectra(r"NAM-90p-10s-5ac"),
    '90%, 10s, 3acc': load_folder_spectra(r"NAM-90p-10s-3ac"),
    '90%, 60s, 5acc': load_folder_spectra(r"90p-60s-5a-samespot"),
    '70%, 60s, 5acc': load_folder_spectra(r"70p-60s-5ac-diifrentpoint"),
}

fig, ax = plt.subplots(figsize=(12, 6))
colors = plt.cm.tab10(np.linspace(0, 1, len(conditions)))
offset = 0

for i, (label, spectra) in enumerate(conditions.items()):
    rs = spectra[0]['raman_shift']
    mask = (rs >= xmin) & (rs <= xmax)
    rs_masked = rs[mask]
    intensities = np.array([sp['intensity'][mask] for sp in spectra])
    mean_intensity = intensities.mean(axis=0)
    ax.plot(rs_masked, mean_intensity + offset, linewidth=1.0,
            color=colors[i], label=f"{label} (n={len(spectra)})")
    offset += mean_intensity.max() * 0.3

ax.set_xlabel('Raman Shift (cm$^{-1}$)', fontsize=12)
ax.set_ylabel('Intensity (a.u.) - stacked', fontsize=12)
ax.set_title('Mean Raman Spectrum per Measurement Condition', fontsize=13, fontweight='bold')
ax.legend(loc='upper right', fontsize=9, framealpha=0.9)
ax.set_xlim(xmin, xmax)

plt.tight_layout()
plt.savefig(r'Analysis\Figures\Figure_3_Conditions_Comparison.png', dpi=300, facecolor='white')
plt.savefig(r'Analysis\Figures\Figure_3_Conditions_Comparison.pdf', facecolor='white')
plt.close()
print("  [OK] Saved.")

# ===== FIGURE 4: SNR Comparison =====
print("\n[4/5] Figure 4: SNR Bar Chart...")

def compute_snr(raman_shift, intensity, signal_range=(900, 1100), noise_range=(1700, 1800)):
    sig_mask = (raman_shift >= signal_range[0]) & (raman_shift <= signal_range[1])
    noise_mask = (raman_shift >= noise_range[0]) & (raman_shift <= noise_range[1])
    signal = np.max(intensity[sig_mask]) - np.median(intensity[noise_mask])
    noise = np.std(intensity[noise_mask])
    return signal / noise if noise > 0 else 0

snr_results = {}
for label, spectra in conditions.items():
    snrs = [compute_snr(sp['raman_shift'], sp['intensity']) for sp in spectra]
    snr_results[label] = snrs

fig, ax = plt.subplots(figsize=(8, 5))
labels = list(snr_results.keys())
means = [np.mean(v) for v in snr_results.values()]
stds = [np.std(v) for v in snr_results.values()]

bars = ax.bar(range(len(labels)), means, yerr=stds, capsize=5,
              color=plt.cm.viridis(np.linspace(0.2, 0.8, len(labels))),
              edgecolor='black', linewidth=0.8, alpha=0.85)

ax.set_xticks(range(len(labels)))
ax.set_xticklabels(labels, rotation=30, ha='right', fontsize=10)
ax.set_ylabel('Signal-to-Noise Ratio', fontsize=12)
ax.set_title('SNR Comparison Across Measurement Conditions', fontsize=13, fontweight='bold')
ax.grid(axis='y', alpha=0.3, linestyle='--')

for bar, mean in zip(bars, means):
    ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 2,
            f'{mean:.0f}', ha='center', va='bottom', fontsize=9, fontweight='bold')

plt.tight_layout()
plt.savefig(r'Analysis\Figures\Figure_4_SNR_Comparison.png', dpi=300, facecolor='white')
plt.savefig(r'Analysis\Figures\Figure_4_SNR_Comparison.pdf', facecolor='white')
plt.close()
print("  [OK] Saved.")

print("\n  SNR Results:")
print(f"  {'Condition':<20} {'Mean SNR':<10} {'Std':<10}")
print("  " + "-" * 40)
for label in labels:
    print(f"  {label:<20} {np.mean(snr_results[label]):<10.1f} {np.std(snr_results[label]):<10.1f}")

# ===== FIGURE 5: Best Spectrum with Peaks =====
print("\n[5/5] Figure 5: Best Spectrum with Peaks...")

best_condition = max(snr_results, key=lambda k: np.max(snr_results[k]))
best_idx = np.argmax(snr_results[best_condition])
best_spectrum = conditions[best_condition][best_idx]

rs = best_spectrum['raman_shift']
intensity = best_spectrum['intensity']
mask = (rs >= xmin) & (rs <= xmax)
rs_plot = rs[mask]
int_plot = intensity[mask]

peaks, properties = find_peaks(int_plot, height=np.median(int_plot) + 2*np.std(int_plot),
                               distance=15, prominence=50)

print(f"  Best: {best_condition} - {best_spectrum['file']}")
print(f"  Peaks: {len(peaks)} detected")
print(f"  Positions (cm-1): {[f'{rs_plot[p]:.0f}' for p in peaks]}")

fig, ax = plt.subplots(figsize=(12, 5))
ax.plot(rs_plot, int_plot, 'k-', linewidth=0.8, label=f'{best_condition}')
ax.plot(rs_plot[peaks], int_plot[peaks], 'rv', markersize=6, label='Detected peaks')

for p in peaks:
    ax.annotate(f'{rs_plot[p]:.0f}',
                xy=(rs_plot[p], int_plot[p]),
                xytext=(0, 10), textcoords='offset points',
                ha='center', va='bottom', fontsize=8, fontweight='bold', color='red')

ax.set_xlabel('Raman Shift (cm$^{-1}$)', fontsize=12)
ax.set_ylabel('Intensity (a.u.)', fontsize=12)
ax.set_title(f'Best Spectrum - {best_condition} ({best_spectrum["file"]})',
             fontsize=13, fontweight='bold')
ax.legend(loc='upper right', fontsize=10)
ax.set_xlim(xmin, xmax)

plt.tight_layout()
plt.savefig(r'Analysis\Figures\Figure_5_Best_Spectrum_Peaks.png', dpi=300, facecolor='white')
plt.savefig(r'Analysis\Figures\Figure_5_Best_Spectrum_Peaks.pdf', facecolor='white')
plt.close()
print("  [OK] Saved.")

print("\n" + "=" * 80)
print("ALL FIGURES GENERATED")
print("=" * 80)
print(f"\nOutput directory: Analysis/Figures/")
for f in sorted(os.listdir(r"Analysis\Figures")):
    size = os.path.getsize(os.path.join(r"Analysis\Figures", f))
    print(f"  {f} ({size/1024:.1f} KB)")
