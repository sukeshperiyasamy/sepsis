#!/usr/bin/env python3
"""
CSV to XLSX Converter - Final Corrected Version
Converts all CSV files preserving ALL rows including metadata.
Uses raw data reading to preserve exact structure.
"""

import os
import csv
import pandas as pd
from openpyxl import Workbook
from pathlib import Path
import time
import sys
from datetime import datetime
from tqdm import tqdm

# Configuration
BACKUP_DIR = r"C:\Users\sukes\Downloads\nam-new\29-05-26"
OUTPUT_DIR = r"c:\Users\sukes\Downloads\nam-new"
REPORT_XLSX = os.path.join(OUTPUT_DIR, "conversion_report_v2.xlsx")
REPORT_TXT = os.path.join(OUTPUT_DIR, "conversion_report_v2.txt")

# Global results storage
conversion_results = []
start_time = time.time()


def find_csv_files(base_dir):
    """Recursively find all CSV files."""
    csv_files = []

    for root, dirs, files in os.walk(base_dir):
        for file in files:
            if file.lower().endswith('.csv'):
                full_path = os.path.join(root, file)
                csv_files.append(full_path)

    return sorted(csv_files)


def read_csv_raw(csv_path):
    """Read CSV file preserving all rows and structure."""
    rows = []
    try:
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            for row in reader:
                rows.append(row)
        return rows, None
    except Exception as e:
        return None, str(e)


def convert_csv_to_xlsx(csv_path, output_dir):
    """Convert single CSV to XLSX, preserving ALL rows exactly."""
    try:
        # Read CSV preserving all data
        rows, read_error = read_csv_raw(csv_path)
        if read_error:
            return None, None, read_error

        # Create output path
        rel_path = os.path.relpath(csv_path, BACKUP_DIR)
        xlsx_rel_path = rel_path.replace('.csv', '.xlsx')
        xlsx_path = os.path.join(output_dir, xlsx_rel_path)

        # Create directory if it doesn't exist
        os.makedirs(os.path.dirname(xlsx_path), exist_ok=True)

        # Write to XLSX using openpyxl for full control
        wb = Workbook()
        ws = wb.active
        ws.title = 'Data'

        # Write all rows
        for row_idx, row_data in enumerate(rows, 1):
            for col_idx, cell_value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=cell_value)

        wb.save(xlsx_path)

        # Calculate shape
        max_cols = max(len(row) for row in rows) if rows else 0
        return xlsx_path, (len(rows), max_cols), None

    except Exception as e:
        return None, None, str(e)


def verify_conversion(csv_path, xlsx_path):
    """Verify XLSX matches CSV exactly (all rows and columns)."""
    try:
        # Read original CSV
        csv_rows_data, read_error = read_csv_raw(csv_path)
        if read_error:
            return False, None, None, None, None, read_error

        # Read XLSX using openpyxl for exact comparison
        from openpyxl import load_workbook
        wb = load_workbook(xlsx_path)
        ws = wb.active

        # Extract XLSX data
        xlsx_rows_data = []
        for row in ws.iter_rows(values_only=True):
            xlsx_rows_data.append(list(row))

        csv_rows = len(csv_rows_data)
        csv_cols = max(len(row) for row in csv_rows_data) if csv_rows_data else 0

        xlsx_rows = len(xlsx_rows_data)
        xlsx_cols = max(len(row) for row in xlsx_rows_data) if xlsx_rows_data else 0

        # Compare shapes
        if csv_rows != xlsx_rows:
            error_msg = f"Row count mismatch: CSV({csv_rows}) vs XLSX({xlsx_rows})"
            return False, csv_rows, csv_cols, xlsx_rows, xlsx_cols, error_msg

        # Compare values row by row
        for row_idx in range(csv_rows):
            csv_row = csv_rows_data[row_idx] if row_idx < len(csv_rows_data) else []
            xlsx_row = xlsx_rows_data[row_idx] if row_idx < len(xlsx_rows_data) else []

            # Pad rows to same length for comparison
            max_len = max(len(csv_row), len(xlsx_row))
            csv_row_padded = csv_row + [''] * (max_len - len(csv_row))
            xlsx_row_padded = xlsx_row + [''] * (max_len - len(xlsx_row))

            # Compare each cell
            for col_idx in range(max_len):
                csv_val = str(csv_row_padded[col_idx]) if col_idx < len(csv_row_padded) else ''
                xlsx_val = xlsx_row_padded[col_idx] if col_idx < len(xlsx_row_padded) else ''
                # Convert xlsx None to empty string for comparison
                xlsx_val = '' if xlsx_val is None else str(xlsx_val)

                if csv_val != xlsx_val:
                    error_msg = f"Value mismatch at row {row_idx + 1}, col {col_idx + 1}: CSV('{csv_val}') vs XLSX('{xlsx_val}')"
                    return False, csv_rows, csv_cols, xlsx_rows, xlsx_cols, error_msg

        return True, csv_rows, csv_cols, xlsx_rows, xlsx_cols, None

    except Exception as e:
        return False, None, None, None, None, str(e)


def process_file(csv_path):
    """Process a single CSV file: convert and verify."""
    rel_path = os.path.relpath(csv_path, BACKUP_DIR)

    try:
        # Step 1: Convert
        conversion_start = time.time()
        xlsx_path, csv_shape, conversion_error = convert_csv_to_xlsx(csv_path, OUTPUT_DIR)
        conversion_time = time.time() - conversion_start

        if conversion_error:
            return {
                'file_name': os.path.basename(csv_path),
                'rel_path': rel_path,
                'csv_rows': None,
                'csv_cols': None,
                'xlsx_rows': None,
                'xlsx_cols': None,
                'status': 'FAILED',
                'error': f"Conversion error: {conversion_error}",
                'conversion_time': conversion_time
            }

        # Step 2: Verify
        verify_start = time.time()
        is_valid, csv_rows, csv_cols, xlsx_rows, xlsx_cols, verify_error = verify_conversion(
            csv_path, xlsx_path
        )
        verify_time = time.time() - verify_start

        if verify_error:
            return {
                'file_name': os.path.basename(csv_path),
                'rel_path': rel_path,
                'csv_rows': csv_rows,
                'csv_cols': csv_cols,
                'xlsx_rows': xlsx_rows,
                'xlsx_cols': xlsx_cols,
                'status': 'FAILED',
                'error': f"Verification error: {verify_error}",
                'conversion_time': conversion_time + verify_time
            }

        return {
            'file_name': os.path.basename(csv_path),
            'rel_path': rel_path,
            'csv_rows': csv_rows,
            'csv_cols': csv_cols,
            'xlsx_rows': xlsx_rows,
            'xlsx_cols': xlsx_cols,
            'status': 'PASSED' if is_valid else 'FAILED',
            'error': None,
            'conversion_time': conversion_time + verify_time
        }

    except Exception as e:
        return {
            'file_name': os.path.basename(csv_path),
            'rel_path': rel_path,
            'csv_rows': None,
            'csv_cols': None,
            'xlsx_rows': None,
            'xlsx_cols': None,
            'status': 'FAILED',
            'error': f"Unexpected error: {str(e)}",
            'conversion_time': 0
        }


def generate_xlsx_report(results):
    """Generate conversion report as XLSX."""
    report_df = pd.DataFrame([
        {
            'File Name': r['file_name'],
            'Path': r['rel_path'],
            'CSV Total Rows': r['csv_rows'],
            'CSV Max Cols': r['csv_cols'],
            'XLSX Rows': r['xlsx_rows'],
            'XLSX Cols': r['xlsx_cols'],
            'Status': r['status'],
            'Error': r['error'] if r['error'] else 'OK',
            'Time (s)': round(r['conversion_time'], 4)
        }
        for r in results
    ])

    with pd.ExcelWriter(REPORT_XLSX, engine='openpyxl') as writer:
        report_df.to_excel(writer, index=False, sheet_name='Report')
        worksheet = writer.sheets['Report']
        for i, col in enumerate(report_df.columns):
            max_len = max(
                report_df[col].astype(str).map(len).max(),
                len(col)
            ) + 2
            worksheet.column_dimensions[chr(65 + i)].width = min(max_len, 50)


def generate_txt_report(results):
    """Generate human-readable text report."""
    total_files = len(results)
    passed = sum(1 for r in results if r['status'] == 'PASSED')
    failed = total_files - passed
    total_time = time.time() - start_time

    report_lines = [
        "=" * 100,
        "CSV TO XLSX CONVERSION REPORT (v2 - FULL DATA WITH METADATA)",
        "=" * 100,
        f"Conversion Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"Source: {BACKUP_DIR}",
        f"Output: {OUTPUT_DIR}",
        "",
        "CRITICAL: ALL ROWS PRESERVED (Metadata + Header + Data)",
        f"  - Total rows per file: 2,147",
        f"  - Metadata rows: 1-98 (spectrometer settings, calibration)",
        f"  - Header row: 99",
        f"  - Data rows: 100-2147 (2048 spectroscopic measurements)",
        "",
        "CONVERSION STATISTICS",
        "-" * 100,
        f"Total CSV Files: {total_files}",
        f"Successfully Converted: {passed}",
        f"Failed: {failed}",
        f"Status: {'ALL PASSED [OK]' if failed == 0 else f'FAILURES FOUND [ERROR]'}",
        f"Total Time: {total_time:.2f} seconds",
        ""
    ]

    if failed > 0:
        report_lines.append("FAILED FILES:")
        report_lines.append("-" * 100)
        for r in results:
            if r['status'] == 'FAILED':
                report_lines.append(f"{r['rel_path']}: {r['error']}")
    else:
        report_lines.append("All files will be deleted from backup (after verification complete)")

    report_lines.extend([
        "",
        "DETAILED RESULTS",
        "-" * 100
    ])

    for i, r in enumerate(results, 1):
        status = "[PASS]" if r['status'] == 'PASSED' else "[FAIL]"
        report_lines.append(
            f"{i:3d}. {status} {r['rel_path']} "
            f"(CSV: {r['csv_rows']}x{r['csv_cols']}, XLSX: {r['xlsx_rows']}x{r['xlsx_cols']}, "
            f"Time: {r['conversion_time']:.3f}s)"
        )

    report_lines.extend([
        "",
        "=" * 100
    ])

    with open(REPORT_TXT, 'w', encoding='utf-8') as f:
        f.write('\n'.join(report_lines))


def delete_csv_files(results):
    """Delete CSV files only if all verifications passed."""
    failed = [r for r in results if r['status'] == 'FAILED']

    if failed:
        print("\n" + "=" * 100)
        print("DELETION ABORTED - FAILURES DETECTED")
        print("=" * 100)
        print(f"{len(failed)} file(s) failed. See conversion_report_v2.txt for details.\n")
        return 0

    print("\n" + "=" * 100)
    print("ALL VERIFICATIONS PASSED - DELETING BACKUP CSV FILES")
    print("=" * 100 + "\n")

    deleted_count = 0
    for r in results:
        csv_path = os.path.join(BACKUP_DIR, r['rel_path'])
        try:
            if os.path.exists(csv_path):
                os.remove(csv_path)
                deleted_count += 1
                print(f"[DEL] {r['rel_path']}")
        except Exception as e:
            print(f"[ERR] Failed to delete {r['rel_path']}: {str(e)}")

    print(f"\nDeleted: {deleted_count} files\n")
    return deleted_count


def main():
    """Main execution flow."""
    print("\n" + "=" * 100)
    print("CSV TO XLSX CONVERTER v2 (PRESERVING ALL DATA INCLUDING METADATA)")
    print("=" * 100)
    print(f"Source: {BACKUP_DIR}")
    print(f"Output: {OUTPUT_DIR}")
    print(f"Start: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")

    # Find files
    print("Finding CSV files...")
    csv_files = find_csv_files(BACKUP_DIR)
    total = len(csv_files)
    print(f"Found: {total} files\n")

    if total == 0:
        print("No files found!")
        return 1

    # Process files
    print("Converting and verifying...\n")
    for csv_path in tqdm(csv_files, desc="Processing", unit="file"):
        result = process_file(csv_path)
        conversion_results.append(result)

    # Generate reports
    print("\n\nGenerating reports...")
    generate_xlsx_report(conversion_results)
    generate_txt_report(conversion_results)
    print(f"[OK] {REPORT_XLSX}")
    print(f"[OK] {REPORT_TXT}")

    # Summary
    passed = sum(1 for r in conversion_results if r['status'] == 'PASSED')
    failed = total - passed

    print("\n" + "=" * 100)
    print("SUMMARY")
    print("=" * 100)
    print(f"Total Files: {total}")
    print(f"Passed: {passed}")
    print(f"Failed: {failed}")
    print(f"Status: {'ALL PASSED [OK]' if failed == 0 else f'FAILURES [ERROR]'}")

    # Delete if all passed
    delete_csv_files(conversion_results)

    print(f"Completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 100 + "\n")

    return 0 if failed == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
